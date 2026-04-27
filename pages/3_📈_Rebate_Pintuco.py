# -*- coding: utf-8 -*-
"""Seguimiento profesional del rebate de proveedores estratégicos."""

import email
import imaplib
import io
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime, timedelta
from email.header import decode_header
from email.utils import parsedate_to_datetime
from typing import Any

import dropbox
import gspread
import pandas as pd
import pytz
import streamlit as st
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from common.treasury_core import (
	INVOICE_EXCLUSION_COLUMNS,
	deactivate_invoice_exclusion,
	load_invoice_exclusion_df,
	normalize_supplier_key,
	register_invoice_exclusion,
)


if "password_correct" not in st.session_state:
	st.session_state["password_correct"] = False

if not st.session_state["password_correct"]:
	st.error("🔒 Debes iniciar sesión para acceder a esta página.")
	st.info("Por favor, ve a la página principal 'Dashboard General' para ingresar la contraseña.")
	st.stop()


st.set_page_config(layout="wide", page_title="Seguimiento Rebate | Proveedores", page_icon="📈")

PINTUCO_ALIASES = ["PINTUCO", "COMPANIA GLOBAL DE PINTURAS"]
PINTUCO_PROVIDER_NAME_ERP = "PINTUCO COLOMBIA S.A.S"
ABRACOL_ALIASES = ["ABRACOL", "ABRASIVOS DE COLOMBIA"]
ABRACOL_PROVIDER_NAMES_ERP = ["ABRACOL S.A.S"]
GOYA_ALIASES = ["GOYA", "GOYAINCOL"]
GOYA_PROVIDER_NAMES_ERP = ["INDUSTRIAS GOYAINCOL SAS", "INDUSTRIAS GOYAINCOL LTDA"]
COLOMBIA_TZ = pytz.timezone("America/Bogota")
CURRENT_CYCLE_NAME = "Ciclo vigente desde el 1 de abril de 2026"
CURRENT_CYCLE_START = date(2026, 4, 1)
FULL_YEAR_CYCLE_NAME = "Ciclo comercial 2026"
FULL_YEAR_CYCLE_START = date(2026, 1, 1)
EXCLUDED_PURCHASE_PERCENT = 0.12
APPLICABLE_PURCHASE_FACTOR = 1 - EXCLUDED_PURCHASE_PERCENT

IMAP_SERVER = "imap.gmail.com"
EMAIL_FOLDER = "TFHKA/Recepcion/Descargados"
DROPBOX_FILE_PATH = "/data/Proveedores.csv"
PINTUCO_WORKSHEET_NAME = "Rebate_Pintuco"
ABRACOL_WORKSHEET_NAME = "Rebate_Abracol"
GOYA_WORKSHEET_NAME = "Rebate_Goya"

MONTHLY_BUDGETS = [
	{"Mes": "Abril", "Mes_Num": 4, "Trimestre": "Q2", "Escala 1": 1456867389.0, "Escala 2": 1544279432.0},
	{"Mes": "Mayo", "Mes_Num": 5, "Trimestre": "Q2", "Escala 1": 2094162232.0, "Escala 2": 2219811966.0},
	{"Mes": "Junio", "Mes_Num": 6, "Trimestre": "Q2", "Escala 1": 2237825409.0, "Escala 2": 2372094934.0},
	{"Mes": "Julio", "Mes_Num": 7, "Trimestre": "Q3", "Escala 1": 2000572886.0, "Escala 2": 2120607260.0},
	{"Mes": "Agosto", "Mes_Num": 8, "Trimestre": "Q3", "Escala 1": 2271787723.0, "Escala 2": 2408094987.0},
	{"Mes": "Septiembre", "Mes_Num": 9, "Trimestre": "Q3", "Escala 1": 2232381407.0, "Escala 2": 2366324292.0},
	{"Mes": "Octubre", "Mes_Num": 10, "Trimestre": "Q4", "Escala 1": 1605138135.0, "Escala 2": 1701446423.0},
	{"Mes": "Noviembre", "Mes_Num": 11, "Trimestre": "Q4", "Escala 1": 1147826895.0, "Escala 2": 1216696508.0},
	{"Mes": "Diciembre", "Mes_Num": 12, "Trimestre": "Q4", "Escala 1": 1555236709.0, "Escala 2": 1648550912.0},
]

MONTHLY_REBATE_RATES = {"Escala 1": 0.01, "Escala 2": 0.015, "Sin escala": 0.0}
QUARTERLY_REBATE_RATES = {"Escala 1": 0.01, "Escala 2": 0.025, "Sin escala": 0.0}
SEASONALITY_TARGET_FACTOR = 0.90
SEASONALITY_RATE = 0.01
CYCLE_RECOMPOSITION_FACTOR = 0.85
SEASONALITY_CUTOFF_OVERRIDES = {"2026-04": date(2026, 4, 24)}

ABRACOL_REBATE_RATE = 0.06
VAT_RATE = 0.19
GOYA_REBATE_NOTE = "La tabla de porcentajes oficial del acuerdo define 40% = 3.5%. La tabla de liquidación semestral muestra 3.0%; el tablero usa la tabla específica de porcentajes como fuente principal."

ABRACOL_BIMESTER_BUDGETS = [
	{"Periodo": "BIMESTRE 1", "Inicio": date(2026, 1, 1), "Fin": date(2026, 2, 28), "Ventas_2025": 168593107.0, "Meta_2026": 185908415.0},
	{"Periodo": "BIMESTRE 2", "Inicio": date(2026, 3, 1), "Fin": date(2026, 4, 30), "Ventas_2025": 161386193.0, "Meta_2026": 188109304.0},
	{"Periodo": "BIMESTRE 3", "Inicio": date(2026, 5, 1), "Fin": date(2026, 6, 30), "Ventas_2025": 171088205.0, "Meta_2026": 202209913.0},
	{"Periodo": "BIMESTRE 4", "Inicio": date(2026, 7, 1), "Fin": date(2026, 8, 31), "Ventas_2025": 151645921.0, "Meta_2026": 193664256.0},
	{"Periodo": "BIMESTRE 5", "Inicio": date(2026, 9, 1), "Fin": date(2026, 10, 31), "Ventas_2025": 194956130.0, "Meta_2026": 196577548.0},
	{"Periodo": "BIMESTRE 6", "Inicio": date(2026, 11, 1), "Fin": date(2026, 12, 31), "Ventas_2025": 159359129.0, "Meta_2026": 191613553.0},
]

GOYA_SEMESTER_BUDGETS = [
	{
		"Periodo": "I SEMESTRE",
		"Inicio": date(2026, 1, 1),
		"Fin": date(2026, 6, 30),
		"Ventas_2024": 206003279.0,
		"Base_2025": 314805149.0,
		"Meta_20": 377766179.0,
		"Meta_30": 409246694.0,
		"Meta_40": 440727209.0,
		"Meta_50": 472207724.0,
	},
	{
		"Periodo": "II SEMESTRE",
		"Inicio": date(2026, 7, 1),
		"Fin": date(2026, 12, 31),
		"Ventas_2024": 222481209.0,
		"Base_2025": 354615080.0,
		"Meta_20": 425538096.0,
		"Meta_30": 460999604.0,
		"Meta_40": 496461112.0,
		"Meta_50": 531922620.0,
	},
]

GOYA_GROWTH_TIERS = [
	{"growth": 0.50, "rate": 0.04, "label": "50%"},
	{"growth": 0.40, "rate": 0.035, "label": "40%"},
	{"growth": 0.30, "rate": 0.025, "label": "30%"},
	{"growth": 0.20, "rate": 0.02, "label": "20%"},
]

INVOICE_COLUMNS = [
	"Fecha_Factura",
	"Numero_Factura",
	"Valor_Neto",
	"Proveedor_Correo",
	"Fecha_Recepcion_Correo",
	"Remitente_Correo",
	"Asunto_Correo",
	"Nombre_Adjunto",
	"Message_ID",
	"Estado_Pago",
]

PROVIDER_EXCLUSION_DISPLAY_COLUMNS = [
	"Fecha_Factura",
	"Numero_Factura",
	"Valor_Neto",
	"Estado_Pago",
	"Fecha_Recepcion_Correo",
	"Remitente_Correo",
	"Asunto_Correo",
	"Nombre_Adjunto",
	"Message_ID",
	"Excluir_De_Calculos",
	"Motivo_Exclusion",
	"Fecha_Exclusion",
]

st.markdown(
	f"""
	<style>
		[data-testid="stSidebar"] {{
			background: linear-gradient(180deg, #0d1c30 0%, #133251 55%, #1c4e80 100%);
			border-right: 1px solid rgba(255,255,255,0.08);
		}}
		[data-testid="stSidebar"] * {{
			color: #f3f7fb;
		}}
		.main .block-container {{
			padding-top: 1rem;
			padding-bottom: 2.4rem;
		}}
		.pintuco-banner {{
			background: radial-gradient(circle at top right, rgba(255,255,255,0.18), transparent 30%), linear-gradient(135deg, #0b2440 0%, #145374 48%, #f0ad1f 100%);
			border-radius: 24px;
			padding: 28px 32px;
			color: #ffffff;
			margin-bottom: 18px;
			box-shadow: 0 18px 42px rgba(11, 36, 64, 0.22);
		}}
		.pintuco-banner h1 {{
			margin: 0;
			font-size: 2.2rem;
		}}
		.pintuco-banner p {{
			margin: 10px 0 0 0;
			font-size: 1rem;
			opacity: 0.95;
		}}
		.info-card {{
			background: linear-gradient(180deg, #f9fbfd 0%, #eef3f8 100%);
			border: 1px solid rgba(12, 45, 87, 0.10);
			border-radius: 18px;
			padding: 16px 18px;
			margin-bottom: 16px;
			box-shadow: 0 10px 24px rgba(12, 45, 87, 0.06);
		}}
		.note-card {{
			background: #fff8e7;
			border: 1px solid rgba(240, 173, 31, 0.28);
			border-radius: 18px;
			padding: 14px 16px;
			margin-bottom: 16px;
		}}
		.kpi-grid {{
			display: grid;
			grid-template-columns: repeat(4, minmax(0, 1fr));
			gap: 12px;
			margin: 12px 0 14px 0;
		}}
		.kpi-card {{
			background: #ffffff;
			border-radius: 20px;
			padding: 16px 18px;
			border: 1px solid rgba(12, 45, 87, 0.08);
			box-shadow: 0 12px 26px rgba(12, 45, 87, 0.06);
		}}
		.kpi-card.navy {{ border-top: 4px solid #0c2d57; }}
		.kpi-card.gold {{ border-top: 4px solid #f0ad1f; }}
		.kpi-card.green {{ border-top: 4px solid #119c63; }}
		.kpi-card.red {{ border-top: 4px solid #d94a4a; }}
		.kpi-label {{
			font-size: 0.72rem;
			text-transform: uppercase;
			letter-spacing: 0.08em;
			color: #5d6c7d;
			margin-bottom: 0.2rem;
		}}
		.kpi-value {{
			font-size: 1.55rem;
			font-weight: 800;
			color: #0c2d57;
			line-height: 1.05;
		}}
		.kpi-sub {{
			font-size: 0.80rem;
			color: #6a7b8f;
			margin-top: 0.28rem;
			line-height: 1.4;
		}}
		.section-card {{
			background: #ffffff;
			border-radius: 22px;
			border: 1px solid rgba(12, 45, 87, 0.08);
			padding: 18px 20px;
			box-shadow: 0 12px 28px rgba(12, 45, 87, 0.06);
			margin-bottom: 16px;
		}}
		.pill {{
			display: inline-block;
			padding: 4px 10px;
			border-radius: 999px;
			font-size: 0.78rem;
			font-weight: 700;
		}}
		.pill.green {{ background: rgba(17, 156, 99, 0.14); color: #0a774a; }}
		.pill.gold {{ background: rgba(240, 173, 31, 0.16); color: #8b6509; }}
		.pill.red {{ background: rgba(217, 74, 74, 0.14); color: #a92f2f; }}
		.pill.navy {{ background: rgba(12, 45, 87, 0.10); color: #0c2d57; }}
		@media (max-width: 1100px) {{
			.kpi-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
		}}
		@media (max-width: 680px) {{
			.kpi-grid {{ grid-template-columns: 1fr; }}
		}}
	</style>
	""",
	unsafe_allow_html=True,
)


def format_currency(value: float) -> str:
	return f"${value:,.0f}"


def format_percent(value: float, decimals: int = 1) -> str:
	return f"{value * 100:.{decimals}f}%"


def safe_divide(numerator: float, denominator: float) -> float:
	if not denominator:
		return 0.0
	return float(numerator) / float(denominator)


def kpi_card_html(label: str, value: str, subtext: str = "", tone: str = "navy") -> str:
	sub_html = f'<div class="kpi-sub">{subtext}</div>' if subtext else ""
	return f'<div class="kpi-card {tone}"><div class="kpi-label">{label}</div><div class="kpi-value">{value}</div>{sub_html}</div>'


def render_kpi_grid(cards: list[str]) -> None:
	st.markdown(f'<div class="kpi-grid">{"".join(cards)}</div>', unsafe_allow_html=True)


def pill_html(text: str, tone: str = "navy") -> str:
	return f'<span class="pill {tone}">{text}</span>'


def normalize_invoice_number(inv_num: Any) -> str:
	if not isinstance(inv_num, str):
		inv_num = str(inv_num)
	return re.sub(r"[^A-Z0-9]", "", inv_num.upper()).strip()


def clean_and_convert_numeric(value: Any) -> float:
	if pd.isna(value) or value is None:
		return 0.0
	cleaned_str = str(value).strip().replace("$", "").replace(",", "")
	try:
		return float(cleaned_str)
	except (ValueError, TypeError):
		return 0.0


def decode_mime_text(value: Any) -> str:
	if not value:
		return ""
	decoded_parts = []
	for part, encoding in decode_header(value):
		if isinstance(part, bytes):
			decoded_parts.append(part.decode(encoding or "utf-8", errors="ignore"))
		else:
			decoded_parts.append(part)
	return "".join(decoded_parts).strip()


def parse_email_datetime(value: str) -> pd.Timestamp:
	if not value:
		return pd.NaT
	try:
		dt = parsedate_to_datetime(value)
		if dt.tzinfo is None:
			dt = COLOMBIA_TZ.localize(dt)
		else:
			dt = dt.astimezone(COLOMBIA_TZ)
		return pd.Timestamp(dt)
	except Exception:
		return pd.NaT


def normalize_local_datetime(value: Any) -> pd.Timestamp:
	if value is None or value == "" or pd.isna(value):
		return pd.NaT

	try:
		timestamp = pd.Timestamp(value)
	except Exception:
		return pd.NaT

	if pd.isna(timestamp):
		return pd.NaT

	try:
		if timestamp.tzinfo is None:
			timestamp = timestamp.tz_localize(COLOMBIA_TZ)
		else:
			timestamp = timestamp.tz_convert(COLOMBIA_TZ)
	except (TypeError, ValueError, AttributeError):
		return pd.NaT

	return timestamp.tz_localize(None)


def normalize_datetime_series(series: pd.Series) -> pd.Series:
	normalized = series.apply(normalize_local_datetime)
	return pd.to_datetime(normalized, errors="coerce")


def sort_invoice_dataframe(df: pd.DataFrame, by: list[str], ascending: bool | list[bool] = True) -> pd.DataFrame:
	if df.empty:
		return df

	sorted_df = df.copy()
	for column in ["Fecha_Factura", "Fecha_Recepcion_Correo"]:
		if column in sorted_df.columns:
			sorted_df[column] = normalize_datetime_series(sorted_df[column])

	return sorted_df.sort_values(by=by, ascending=ascending, na_position="last")


def build_abracol_budget_frame() -> pd.DataFrame:
	return pd.DataFrame(ABRACOL_BIMESTER_BUDGETS).copy()


def build_goya_budget_frame() -> pd.DataFrame:
	return pd.DataFrame(GOYA_SEMESTER_BUDGETS).copy()


def get_provider_configs() -> dict[str, dict[str, Any]]:
	return {
		"pintuco": {
			"key": "pintuco",
			"label": "Pintuco",
			"title": "Dashboard Ejecutivo de Rebate Pintuco",
			"cycle_name": CURRENT_CYCLE_NAME,
			"cycle_start": CURRENT_CYCLE_START,
			"aliases": PINTUCO_ALIASES,
			"provider_names_erp": [PINTUCO_PROVIDER_NAME_ERP],
			"worksheet_name": PINTUCO_WORKSHEET_NAME,
			"excluded_purchase_percent": EXCLUDED_PURCHASE_PERCENT,
		},
		"abracol": {
			"key": "abracol",
			"label": "Abracol",
			"title": "Dashboard Ejecutivo de Rebate Abracol",
			"cycle_name": FULL_YEAR_CYCLE_NAME,
			"cycle_start": FULL_YEAR_CYCLE_START,
			"aliases": ABRACOL_ALIASES,
			"provider_names_erp": ABRACOL_PROVIDER_NAMES_ERP,
			"worksheet_name": ABRACOL_WORKSHEET_NAME,
			"excluded_purchase_percent": 0.0,
			"flat_rate": ABRACOL_REBATE_RATE,
			"budget_df": build_abracol_budget_frame(),
		},
		"goya": {
			"key": "goya",
			"label": "Goya",
			"title": "Dashboard Ejecutivo de Rebate Goya",
			"cycle_name": FULL_YEAR_CYCLE_NAME,
			"cycle_start": FULL_YEAR_CYCLE_START,
			"aliases": GOYA_ALIASES,
			"provider_names_erp": GOYA_PROVIDER_NAMES_ERP,
			"worksheet_name": GOYA_WORKSHEET_NAME,
			"excluded_purchase_percent": 0.0,
			"budget_df": build_goya_budget_frame(),
			"growth_tiers": GOYA_GROWTH_TIERS,
			"note": GOYA_REBATE_NOTE,
		},
	}


def get_third_sunday(month_start: pd.Timestamp) -> date:
	first_day = month_start.date().replace(day=1)
	days_until_sunday = (6 - first_day.weekday()) % 7
	first_sunday = first_day + timedelta(days=days_until_sunday)
	return first_sunday + timedelta(days=14)


def build_budget_frame() -> pd.DataFrame:
	budget_df = pd.DataFrame(MONTHLY_BUDGETS).copy()
	budget_df["Mes_Inicio"] = pd.to_datetime({"year": CURRENT_CYCLE_START.year, "month": budget_df["Mes_Num"], "day": 1})
	budget_df["Mes_Clave"] = budget_df["Mes_Inicio"].dt.strftime("%Y-%m")
	budget_df["Corte_Estacionalidad"] = budget_df.apply(
		lambda row: pd.Timestamp(SEASONALITY_CUTOFF_OVERRIDES.get(row["Mes_Clave"], get_third_sunday(row["Mes_Inicio"]))),
		axis=1,
	)
	return budget_df


def get_rebate_configuration() -> dict:
	budget_df = build_budget_frame()
	with st.sidebar:
		st.header("Motor comercial")
		st.caption("Parámetros fijos del acuerdo comercial. Sin edición manual en pantalla para evitar lecturas ambiguas.")
		st.markdown(
			"\n".join(
				[
					f"- Rebate mensual: Escala 1 {format_percent(MONTHLY_REBATE_RATES['Escala 1'], 1)} | Escala 2 {format_percent(MONTHLY_REBATE_RATES['Escala 2'], 1)}",
					f"- Rebate trimestral: Escala 1 {format_percent(QUARTERLY_REBATE_RATES['Escala 1'], 1)} | Escala 2 {format_percent(QUARTERLY_REBATE_RATES['Escala 2'], 1)}",
					f"- Estacionalidad: {format_percent(SEASONALITY_TARGET_FACTOR, 0)} de Escala 2 y bono de {format_percent(SEASONALITY_RATE, 1)}",
					f"- Recomposición 9 meses: {format_percent(CYCLE_RECOMPOSITION_FACTOR, 0)} del saldo elegible",
					"- Abril 2026: corte especial de estacionalidad extendido al 2026-04-24",
				]
			)
		)

	return {
		"budget_df": budget_df,
		"monthly_rates": MONTHLY_REBATE_RATES.copy(),
		"quarterly_rates": QUARTERLY_REBATE_RATES.copy(),
		"seasonality_target_factor": SEASONALITY_TARGET_FACTOR,
		"seasonality_rate": SEASONALITY_RATE,
		"cycle_recomposition_factor": CYCLE_RECOMPOSITION_FACTOR,
	}


@st.cache_resource(show_spinner="Conectando a Google Sheets...")
def connect_to_google_sheets():
	try:
		scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
		creds = Credentials.from_service_account_info(st.secrets["google_credentials"], scopes=scopes)
		return gspread.authorize(creds)
	except Exception as exc:
		st.error(f"❌ Error crítico al autenticar con Google: {exc}")
		return None


def get_worksheet(client: gspread.Client, sheet_key: str, worksheet_name: str):
	try:
		spreadsheet = client.open_by_key(sheet_key)
	except gspread.exceptions.APIError:
		st.error("❌ No se pudo abrir el libro de Google Sheets. Verifica el ID y el acceso del correo de servicio.")
		st.info(f"Correo de servicio: {st.secrets.google_credentials['client_email']}")
		st.stop()
		return None
	try:
		return spreadsheet.worksheet(worksheet_name)
	except gspread.WorksheetNotFound:
		st.warning(f"La pestaña '{worksheet_name}' no existe. Se creará automáticamente.")
		return spreadsheet.add_worksheet(title=worksheet_name, rows="2000", cols="50")


def update_gsheet_from_df(worksheet: gspread.Worksheet, df: pd.DataFrame) -> bool:
	try:
		df_to_upload = df.copy()
		for column in ["Fecha_Factura", "Fecha_Recepcion_Correo"]:
			if column in df_to_upload.columns:
				df_to_upload[column] = pd.to_datetime(df_to_upload[column], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")

		df_to_upload = df_to_upload.astype(str).replace({"nan": "", "NaT": "", "None": ""})
		worksheet.clear()
		worksheet.update([df_to_upload.columns.tolist()] + df_to_upload.values.tolist(), "A1")
		return True
	except Exception as exc:
		st.error(f"❌ Error al actualizar la hoja '{worksheet.title}': {exc}")
		return False


@st.cache_data(ttl=600, show_spinner="Descargando cartera vigente de Dropbox...")
def load_pending_documents_from_dropbox(provider_names_erp: tuple[str, ...]) -> set:
	try:
		dbx = dropbox.Dropbox(
			oauth2_refresh_token=st.secrets.dropbox["refresh_token"],
			app_key=st.secrets.dropbox["app_key"],
			app_secret=st.secrets.dropbox["app_secret"],
		)
		_, response = dbx.files_download(DROPBOX_FILE_PATH)
		df = pd.read_csv(
			io.StringIO(response.content.decode("latin1")),
			sep="{",
			header=None,
			engine="python",
			names=[
				"nombre_proveedor_erp",
				"serie",
				"num_entrada",
				"num_factura",
				"doc_erp",
				"fecha_emision_erp",
				"fecha_vencimiento_erp",
				"valor_total_erp",
			],
		)
		provider_filter = {provider.upper() for provider in provider_names_erp}
		provider_df = df[df["nombre_proveedor_erp"].fillna("").astype(str).str.upper().isin(provider_filter)].copy()
		provider_df["valor_total_erp"] = provider_df["valor_total_erp"].apply(clean_and_convert_numeric)

		credit_note_mask = (provider_df["valor_total_erp"] < 0) & (
			provider_df["num_factura"].isna() | (provider_df["num_factura"].astype(str).str.strip() == "")
		)
		if credit_note_mask.any():
			provider_df.loc[credit_note_mask, "num_factura"] = (
				"NC-"
				+ provider_df.loc[credit_note_mask, "doc_erp"].astype(str).str.strip()
				+ "-"
				+ provider_df.loc[credit_note_mask, "valor_total_erp"].abs().astype(int).astype(str)
			)

		provider_df.dropna(subset=["num_factura"], inplace=True)
		return set(provider_df["num_factura"].apply(normalize_invoice_number))
	except Exception as exc:
		st.error(f"❌ Error cargando cartera de Dropbox: {exc}")
		return set()


def parse_invoice_xml(xml_content: str, aliases: list[str]) -> dict | None:
	try:
		namespaces = {
			"cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
			"cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
		}

		xml_content = re.sub(r"^[^<]+", "", xml_content.strip())
		root = ET.fromstring(xml_content.encode("utf-8"))

		description_node = root.find(".//cac:Attachment/cac:ExternalReference/cbc:Description", namespaces)
		if description_node is not None and description_node.text and "<Invoice" in description_node.text:
			nested_xml = re.sub(r"^[^<]+", "", description_node.text.strip())
			invoice_root = ET.fromstring(nested_xml.encode("utf-8"))
		else:
			invoice_root = root

		supplier_name_node = invoice_root.find(
			".//cac:AccountingSupplierParty/cac:Party/cac:PartyLegalEntity/cbc:RegistrationName",
			namespaces,
		)
		if supplier_name_node is None or not supplier_name_node.text:
			return None

		supplier_name = supplier_name_node.text.strip()
		if not any(alias in supplier_name.upper() for alias in aliases):
			return None

		invoice_number_node = invoice_root.find("./cbc:ID", namespaces)
		issue_date_node = invoice_root.find("./cbc:IssueDate", namespaces)
		net_value_node = invoice_root.find(".//cac:LegalMonetaryTotal/cbc:TaxExclusiveAmount", namespaces)

		if invoice_number_node is None or issue_date_node is None or net_value_node is None:
			return None

		return {
			"Fecha_Factura": issue_date_node.text.strip(),
			"Numero_Factura": normalize_invoice_number(invoice_number_node.text.strip()),
			"Valor_Neto": clean_and_convert_numeric(net_value_node.text.strip()),
			"Proveedor_Correo": supplier_name,
		}
	except Exception:
		return None


def extract_invoice_records_from_message(message_obj, aliases: list[str]) -> tuple[list[dict], dict]:
	records = []
	stats = {"attachments_scanned": 0, "xml_files_scanned": 0, "invoice_rows_detected": 0}

	email_subject = decode_mime_text(message_obj.get("Subject", ""))
	email_sender = decode_mime_text(message_obj.get("From", ""))
	email_message_id = decode_mime_text(message_obj.get("Message-ID", ""))
	email_received_at = parse_email_datetime(message_obj.get("Date", ""))

	for part in message_obj.walk():
		if part.get_content_maintype() == "multipart":
			continue

		filename = decode_mime_text(part.get_filename() or "")
		content_type = (part.get_content_type() or "").lower()
		payload = part.get_payload(decode=True)
		if payload is None:
			continue

		is_zip = filename.lower().endswith(".zip") or content_type in {"application/zip", "application/x-zip-compressed"}
		is_xml = filename.lower().endswith(".xml") or content_type in {"application/xml", "text/xml"}
		if not is_zip and not is_xml:
			continue

		stats["attachments_scanned"] += 1

		if is_zip:
			try:
				with zipfile.ZipFile(io.BytesIO(payload)) as zip_file:
					for internal_name in zip_file.namelist():
						if not internal_name.lower().endswith(".xml"):
							continue
						stats["xml_files_scanned"] += 1
						xml_content = zip_file.read(internal_name).decode("utf-8", "ignore")
						details = parse_invoice_xml(xml_content, aliases)
						if details:
							details.update(
								{
									"Fecha_Recepcion_Correo": email_received_at,
									"Remitente_Correo": email_sender,
									"Asunto_Correo": email_subject,
									"Nombre_Adjunto": internal_name,
									"Message_ID": email_message_id,
								}
							)
							records.append(details)
							stats["invoice_rows_detected"] += 1
			except Exception:
				continue

		if is_xml:
			stats["xml_files_scanned"] += 1
			xml_content = payload.decode("utf-8", "ignore")
			details = parse_invoice_xml(xml_content, aliases)
			if details:
				details.update(
					{
						"Fecha_Recepcion_Correo": email_received_at,
						"Remitente_Correo": email_sender,
						"Asunto_Correo": email_subject,
						"Nombre_Adjunto": filename or "adjunto_xml",
						"Message_ID": email_message_id,
					}
				)
				records.append(details)
				stats["invoice_rows_detected"] += 1

	return records, stats


def fetch_provider_invoices_from_email(start_date: date, aliases: list[str], provider_label: str) -> tuple[pd.DataFrame, dict]:
	invoices_data = []
	stats = {
		"emails_found": 0,
		"emails_processed": 0,
		"attachments_scanned": 0,
		"xml_files_scanned": 0,
		"invoice_rows_detected": 0,
		"started_from": start_date.strftime("%Y-%m-%d"),
	}

	try:
		mail = imaplib.IMAP4_SSL(IMAP_SERVER)
		mail.login(st.secrets.email["address"], st.secrets.email["password"])
		mail.select(f'"{EMAIL_FOLDER}"')

		search_query = f'(SINCE "{start_date.strftime("%d-%b-%Y")}")'
		_, messages = mail.search(None, search_query)
		message_ids = messages[0].split()
		stats["emails_found"] = len(message_ids)

		if not message_ids:
			mail.logout()
			return pd.DataFrame(), stats

		progress_text = f"Procesando {len(message_ids)} correos del buzón de {provider_label}..."
		progress_bar = st.progress(0, text=progress_text)

		for index, message_id in enumerate(message_ids, start=1):
			_, data = mail.fetch(message_id, "(RFC822)")
			message_obj = email.message_from_bytes(data[0][1])
			records, email_stats = extract_invoice_records_from_message(message_obj, aliases)
			invoices_data.extend(records)

			stats["emails_processed"] += 1
			stats["attachments_scanned"] += email_stats["attachments_scanned"]
			stats["xml_files_scanned"] += email_stats["xml_files_scanned"]
			stats["invoice_rows_detected"] += email_stats["invoice_rows_detected"]

			progress_bar.progress(index / len(message_ids), text=f"{progress_text} ({index}/{len(message_ids)})")

		mail.logout()
		return pd.DataFrame(invoices_data), stats
	except Exception as exc:
		st.error(f"❌ Error procesando correos de {provider_label}: {exc}")
		return pd.DataFrame(), stats


def ensure_invoice_columns(df: pd.DataFrame) -> pd.DataFrame:
	normalized_df = df.copy()
	for column in INVOICE_COLUMNS:
		if column not in normalized_df.columns:
			normalized_df[column] = ""
	return normalized_df.reindex(columns=INVOICE_COLUMNS)


def prepare_invoice_dataframe(df: pd.DataFrame, excluded_purchase_percent: float = EXCLUDED_PURCHASE_PERCENT) -> pd.DataFrame:
	if df.empty:
		return ensure_invoice_columns(df)

	prepared = ensure_invoice_columns(df)
	prepared["Fecha_Factura"] = normalize_datetime_series(prepared["Fecha_Factura"])
	prepared["Fecha_Recepcion_Correo"] = normalize_datetime_series(prepared["Fecha_Recepcion_Correo"])
	prepared["Valor_Neto"] = pd.to_numeric(prepared["Valor_Neto"], errors="coerce").fillna(0.0)
	prepared["Numero_Factura"] = prepared["Numero_Factura"].apply(normalize_invoice_number)
	prepared["Compra_Excluida_12"] = prepared["Valor_Neto"] * excluded_purchase_percent
	prepared["Compra_Aplicable_Rebate"] = prepared["Valor_Neto"] * (1 - excluded_purchase_percent)
	return prepared


@st.cache_data(ttl=120)
def load_invoice_exclusion_registry() -> pd.DataFrame:
	gs_client = connect_to_google_sheets()
	if not gs_client:
		return pd.DataFrame(columns=INVOICE_EXCLUSION_COLUMNS)
	return load_invoice_exclusion_df(gs_client)


def build_provider_invoice_key(provider_config: dict[str, Any], invoice_number: Any) -> str:
	provider_norm = normalize_supplier_key(provider_config["provider_names_erp"][0])
	return f"{provider_norm}|{normalize_invoice_number(invoice_number)}"


def apply_provider_invoice_exclusions(provider_df: pd.DataFrame, provider_config: dict[str, Any], exclusion_df: pd.DataFrame) -> pd.DataFrame:
	prepared = provider_df.copy()
	prepared["Invoice_Key"] = prepared["Numero_Factura"].apply(lambda value: build_provider_invoice_key(provider_config, value))
	prepared["Excluir_De_Calculos"] = False
	prepared["Motivo_Exclusion"] = ""
	prepared["Fecha_Exclusion"] = pd.NaT
	prepared["Exclusion_ID"] = ""

	if exclusion_df.empty:
		return prepared

	active_exclusions = exclusion_df[exclusion_df["status"].astype(str).str.upper().eq("ACTIVO")].copy()
	if active_exclusions.empty:
		return prepared

	active_exclusions.sort_values(by="created_at", inplace=True)
	active_exclusions = active_exclusions.drop_duplicates(subset=["invoice_key"], keep="last")
	lookup = active_exclusions.set_index("invoice_key")
	mask = prepared["Invoice_Key"].astype(str).isin(lookup.index)
	if not mask.any():
		return prepared

	prepared.loc[mask, "Excluir_De_Calculos"] = True
	prepared.loc[mask, "Motivo_Exclusion"] = prepared.loc[mask, "Invoice_Key"].map(lookup["reason"]).fillna("")
	prepared.loc[mask, "Fecha_Exclusion"] = prepared.loc[mask, "Invoice_Key"].map(lookup["created_at"])
	prepared.loc[mask, "Exclusion_ID"] = prepared.loc[mask, "Invoice_Key"].map(lookup["exclusion_id"]).fillna("")
	return prepared


def split_provider_dataset(provider_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
	prepared = provider_df.copy()
	prepared["Excluir_De_Calculos"] = prepared.get("Excluir_De_Calculos", pd.Series(False, index=prepared.index)).astype(bool)
	active_df = prepared[~prepared["Excluir_De_Calculos"]].copy()
	excluded_df = prepared[prepared["Excluir_De_Calculos"]].copy()
	return active_df, excluded_df


def get_last_sync_key(provider_config: dict[str, Any]) -> str:
	return f"last_{provider_config['key']}_sync"


def get_last_sync_stats_key(provider_config: dict[str, Any]) -> str:
	return f"last_{provider_config['key']}_sync_stats"


def run_provider_sync(provider_config: dict[str, Any]) -> None:
	provider_label = provider_config["label"]
	cycle_start = provider_config["cycle_start"]
	excluded_purchase_percent = provider_config.get("excluded_purchase_percent", 0.0)
	pending_docs_set = load_pending_documents_from_dropbox(tuple(provider_config["provider_names_erp"]))

	with st.spinner(f"Sincronizando facturas y trazabilidad de {provider_label}..."):
		gs_client = connect_to_google_sheets()
		if not gs_client:
			st.error("Sincronización cancelada. No fue posible conectar con Google Sheets.")
			st.stop()

		worksheet = get_worksheet(gs_client, st.secrets["google_sheet_id"], provider_config["worksheet_name"])
		historical_df = pd.DataFrame(columns=INVOICE_COLUMNS)
		start_date = cycle_start

		try:
			records = worksheet.get_all_records()
			if records:
				historical_df = prepare_invoice_dataframe(pd.DataFrame(records), excluded_purchase_percent)
				historical_df = historical_df[historical_df["Fecha_Factura"].dt.date >= cycle_start].copy()
				if not historical_df.empty:
					last_sync_date = historical_df["Fecha_Factura"].max().date()
					start_date = max(cycle_start, last_sync_date - timedelta(days=3))
		except Exception as exc:
			st.warning(f"No se pudo leer el histórico de Google Sheets. Se sincronizará desde el inicio del ciclo. Detalle: {exc}")

		st.info(f"Buscando facturas desde {start_date.strftime('%Y-%m-%d')} en la carpeta de correo {EMAIL_FOLDER}.")
		new_invoices_df, sync_stats = fetch_provider_invoices_from_email(start_date, provider_config["aliases"], provider_label)

		combined_df = prepare_invoice_dataframe(historical_df.copy(), excluded_purchase_percent)
		if not new_invoices_df.empty:
			new_invoices_df = prepare_invoice_dataframe(new_invoices_df, excluded_purchase_percent)
			combined_df = prepare_invoice_dataframe(pd.concat([historical_df, new_invoices_df], ignore_index=True), excluded_purchase_percent)
			combined_df = sort_invoice_dataframe(combined_df, by=["Fecha_Factura", "Fecha_Recepcion_Correo"])
			combined_df.drop_duplicates(subset=["Numero_Factura"], keep="last", inplace=True)
			st.success(f"Se consolidaron {len(new_invoices_df)} registros de factura detectados desde el correo.")
		else:
			st.info(f"No se detectaron facturas nuevas de {provider_label} en el correo para el rango consultado.")

		if not combined_df.empty:
			combined_df["Estado_Pago"] = combined_df["Numero_Factura"].apply(
				lambda number: "Pendiente" if normalize_invoice_number(number) in pending_docs_set else "Pagada"
			)
			combined_df = ensure_invoice_columns(combined_df)

			if update_gsheet_from_df(worksheet, sort_invoice_dataframe(combined_df, by=["Fecha_Factura", "Numero_Factura"])):
				st.success(f"✅ Base de datos de {provider_label} actualizada correctamente.")
			else:
				st.error("❌ La actualización de Google Sheets falló.")
		else:
			st.warning(f"No hay facturas de {provider_label} para guardar en la hoja.")

		st.session_state[get_last_sync_key(provider_config)] = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M:%S")
		st.session_state[get_last_sync_stats_key(provider_config)] = sync_stats


def fetch_pintuco_invoices_from_email(start_date: date) -> tuple[pd.DataFrame, dict]:
	return fetch_provider_invoices_from_email(start_date, PINTUCO_ALIASES, "Pintuco")


def build_tracking_alerts(df: pd.DataFrame) -> list[str]:
	if df.empty:
		return ["No hay facturas cargadas para evaluar alertas del ciclo."]

	alerts = []
	missing_invoice_date = int(df["Fecha_Factura"].isna().sum())
	missing_email_date = int(df["Fecha_Recepcion_Correo"].isna().sum())
	missing_message_id = int(df["Message_ID"].fillna("").astype(str).str.strip().eq("").sum())
	duplicates = int(df["Numero_Factura"].fillna("").duplicated(keep=False).sum())
	pending_count = int((df["Estado_Pago"] == "Pendiente").sum())

	if missing_invoice_date:
		alerts.append(f"Hay {missing_invoice_date} facturas sin fecha de emisión válida.")
	if missing_email_date:
		alerts.append(f"Hay {missing_email_date} registros sin fecha de recepción del correo.")
	if missing_message_id:
		alerts.append(f"Hay {missing_message_id} registros sin Message-ID; la auditoría del correo queda incompleta.")
	if duplicates:
		alerts.append(f"Se detectaron {duplicates} filas con número de factura repetido en la vista actual.")
	if pending_count:
		alerts.append(f"Siguen pendientes de pago {pending_count} facturas del rango filtrado.")
	if not alerts:
		alerts.append("La trazabilidad está completa para el rango filtrado y no se detectan alertas operativas.")

	return alerts


def run_pintuco_sync():
	run_provider_sync(get_provider_configs()["pintuco"])


@st.cache_data(ttl=300)
def load_provider_data_from_gsheet(worksheet_name: str, cycle_start: date, excluded_purchase_percent: float) -> pd.DataFrame:
	try:
		gs_client = connect_to_google_sheets()
		worksheet = get_worksheet(gs_client, st.secrets["google_sheet_id"], worksheet_name)
		records = worksheet.get_all_records()
		if not records:
			return pd.DataFrame()

		df = prepare_invoice_dataframe(pd.DataFrame(records), excluded_purchase_percent)
		df = df[df["Fecha_Factura"].dt.date >= cycle_start].copy()
		return sort_invoice_dataframe(df, by=["Fecha_Factura", "Numero_Factura"])
	except Exception as exc:
		st.error(f"❌ Error al cargar datos desde Google Sheets: {exc}")
		return pd.DataFrame()


def load_pintuco_data_from_gsheet() -> pd.DataFrame:
	return load_provider_data_from_gsheet(PINTUCO_WORKSHEET_NAME, CURRENT_CYCLE_START, EXCLUDED_PURCHASE_PERCENT)


def build_cycle_summary(df: pd.DataFrame) -> pd.DataFrame:
	if df.empty:
		return pd.DataFrame()

	return df[
		[
			"Mes",
			"Trimestre",
			"Compra_Neta",
			"Exclusion_12",
			"Compra_Aplicable",
			"Presupuesto_Escala_1",
			"Presupuesto_Escala_2",
			"Cumplimiento_E1",
			"Cumplimiento_E2",
			"Escala_Lograda",
			"Rebate_Mensual_Ganado",
			"Bono_Estacionalidad",
			"Pendiente_Cartera",
		]
	].copy()


def determine_scale(actual_purchase: float, scale_1_target: float, scale_2_target: float) -> str:
	if scale_2_target > 0 and actual_purchase >= scale_2_target:
		return "Escala 2"
	if scale_1_target > 0 and actual_purchase >= scale_1_target:
		return "Escala 1"
	return "Sin escala"


def get_scale_tone(scale_name: str) -> str:
	if scale_name == "Escala 2":
		return "green"
	if scale_name == "Escala 1":
		return "gold"
	if any(token in scale_name for token in ["No", "Riesgo", "Bloqueada"]):
		return "red"
	return "navy"


def build_monthly_rebate_table(df: pd.DataFrame, budget_df: pd.DataFrame, config: dict, snapshot_date: date) -> pd.DataFrame:
	working_df = df.copy()
	if not working_df.empty:
		working_df["Mes_Inicio"] = working_df["Fecha_Factura"].dt.to_period("M").dt.to_timestamp()

	rows = []
	for _, budget_row in budget_df.iterrows():
		month_start = budget_row["Mes_Inicio"]
		month_end = (month_start + pd.offsets.MonthEnd(0)).date()
		cutoff = pd.Timestamp(budget_row["Corte_Estacionalidad"])
		month_df = working_df[working_df["Mes_Inicio"] == month_start].copy() if not working_df.empty else pd.DataFrame(columns=df.columns)

		purchase_net = float(month_df["Valor_Neto"].sum()) if not month_df.empty else 0.0
		excluded_12 = float(month_df["Compra_Excluida_12"].sum()) if not month_df.empty else 0.0
		purchase_applicable = float(month_df["Compra_Aplicable_Rebate"].sum()) if not month_df.empty else 0.0
		pending_value = float(month_df.loc[month_df["Estado_Pago"] == "Pendiente", "Valor_Neto"].sum()) if not month_df.empty else 0.0
		cutoff_purchase = float(month_df.loc[month_df["Fecha_Factura"] <= cutoff, "Compra_Aplicable_Rebate"].sum()) if not month_df.empty else 0.0
		invoice_count = int(month_df["Numero_Factura"].nunique()) if not month_df.empty else 0
		pending_count = int((month_df["Estado_Pago"] == "Pendiente").sum()) if not month_df.empty else 0
		paid_count = int((month_df["Estado_Pago"] == "Pagada").sum()) if not month_df.empty else 0

		scale_1_target = float(budget_row["Escala 1"])
		scale_2_target = float(budget_row["Escala 2"])
		monthly_scale = determine_scale(purchase_applicable, scale_1_target, scale_2_target)
		monthly_rate = config["monthly_rates"].get(monthly_scale, 0.0)
		monthly_rebate = purchase_applicable * monthly_rate

		seasonality_target = scale_2_target * config["seasonality_target_factor"]
		seasonality_progress = safe_divide(cutoff_purchase, seasonality_target)
		seasonality_met = seasonality_target > 0 and cutoff_purchase >= seasonality_target
		current_month = snapshot_date.year == month_start.year and snapshot_date.month == month_start.month
		if snapshot_date < month_start.date():
			month_status = "Futuro"
			seasonality_status = "Futuro"
		elif monthly_scale != "Sin escala":
			month_status = "Cumplida"
			seasonality_status = "Cumplida" if seasonality_met else ("En ventana" if current_month and snapshot_date <= cutoff.date() else "No cumplida")
		elif snapshot_date <= month_end:
			month_status = "Abierta"
			seasonality_status = "En ventana" if snapshot_date <= cutoff.date() else "No cumplida"
		else:
			month_status = "No cumplida"
			seasonality_status = "No cumplida"

		rows.append(
			{
				"Mes": budget_row["Mes"],
				"Mes_Clave": budget_row["Mes_Clave"],
				"Mes_Inicio": month_start,
				"Trimestre": budget_row["Trimestre"],
				"Facturas": invoice_count,
				"Pagadas": paid_count,
				"Pendientes": pending_count,
				"Compra_Neta": purchase_net,
				"Exclusion_12": excluded_12,
				"Compra_Aplicable": purchase_applicable,
				"Presupuesto_Escala_1": scale_1_target,
				"Presupuesto_Escala_2": scale_2_target,
				"Cumplimiento_E1": safe_divide(purchase_applicable, scale_1_target),
				"Cumplimiento_E2": safe_divide(purchase_applicable, scale_2_target),
				"Faltante_E1": max(scale_1_target - purchase_applicable, 0.0),
				"Faltante_E2": max(scale_2_target - purchase_applicable, 0.0),
				"Escala_Lograda": monthly_scale,
				"Estado_Mes": month_status,
				"Rebate_Mensual_Pct": monthly_rate,
				"Rebate_Mensual_Ganado": monthly_rebate,
				"Corte_Estacionalidad": cutoff,
				"Compra_Hasta_Corte": cutoff_purchase,
				"Meta_Estacionalidad": seasonality_target,
				"Avance_Estacionalidad": seasonality_progress,
				"Estado_Estacionalidad": seasonality_status,
				"Bono_Estacionalidad": purchase_applicable * config["seasonality_rate"] if seasonality_met else 0.0,
				"Pendiente_Cartera": pending_value,
				"Cartera_Riesgo": pending_value > 0,
				"Mes_Cerrado": snapshot_date > month_end,
			}
		)

	return pd.DataFrame(rows)


def build_quarterly_rebate_table(monthly_df: pd.DataFrame, config: dict) -> pd.DataFrame:
	rows = []
	for quarter_name, quarter_df in monthly_df.groupby("Trimestre", sort=False):
		purchase_applicable = float(quarter_df["Compra_Aplicable"].sum())
		target_e1 = float(quarter_df["Presupuesto_Escala_1"].sum())
		target_e2 = float(quarter_df["Presupuesto_Escala_2"].sum())
		quarter_scale = determine_scale(purchase_applicable, target_e1, target_e2)
		quarter_rate = config["quarterly_rates"].get(quarter_scale, 0.0)
		quarter_rebate = purchase_applicable * quarter_rate

		recomposition_eligible = 0.0
		recomposition_blocked = 0.0
		quarter_month_rate = config["monthly_rates"].get(quarter_scale, 0.0)
		if quarter_scale != "Sin escala":
			for _, month_row in quarter_df.iterrows():
				rate_gap = max(quarter_month_rate - float(month_row["Rebate_Mensual_Pct"]), 0.0)
				recoverable_value = float(month_row["Compra_Aplicable"]) * rate_gap
				if month_row["Cartera_Riesgo"]:
					recomposition_blocked += recoverable_value
				else:
					recomposition_eligible += recoverable_value

		rows.append(
			{
				"Trimestre": quarter_name,
				"Compra_Aplicable": purchase_applicable,
				"Presupuesto_Escala_1": target_e1,
				"Presupuesto_Escala_2": target_e2,
				"Cumplimiento_E1": safe_divide(purchase_applicable, target_e1),
				"Cumplimiento_E2": safe_divide(purchase_applicable, target_e2),
				"Faltante_E1": max(target_e1 - purchase_applicable, 0.0),
				"Faltante_E2": max(target_e2 - purchase_applicable, 0.0),
				"Escala_Lograda": quarter_scale,
				"Rebate_Trimestral_Pct": quarter_rate,
				"Rebate_Trimestral_Ganado": quarter_rebate,
				"Recomposicion_Trimestral_Proyectada": recomposition_eligible,
				"Recomposicion_Cartera_Bloqueada": recomposition_blocked,
				"Meses_Cubiertos": ", ".join(quarter_df["Mes"].tolist()),
			}
		)

	return pd.DataFrame(rows)


def build_cycle_projection(monthly_df: pd.DataFrame, budget_df: pd.DataFrame, config: dict, snapshot_date: date) -> dict:
	current_month_start = pd.Timestamp(snapshot_date.replace(day=1))
	elapsed_monthly = monthly_df[monthly_df["Mes_Inicio"] <= current_month_start].copy()
	elapsed_budget = budget_df[budget_df["Mes_Inicio"] <= current_month_start].copy()

	applicable_purchase = float(elapsed_monthly["Compra_Aplicable"].sum()) if not elapsed_monthly.empty else 0.0
	target_elapsed_e1 = float(elapsed_budget["Escala 1"].sum()) if not elapsed_budget.empty else 0.0
	target_elapsed_e2 = float(elapsed_budget["Escala 2"].sum()) if not elapsed_budget.empty else 0.0
	target_total_e1 = float(budget_df["Escala 1"].sum())
	target_total_e2 = float(budget_df["Escala 2"].sum())
	cycle_scale = determine_scale(applicable_purchase, target_elapsed_e1, target_elapsed_e2)
	cycle_month_rate = config["monthly_rates"].get(cycle_scale, 0.0)

	recoverable_pool = 0.0
	blocked_pool = 0.0
	if cycle_scale != "Sin escala":
		for _, month_row in elapsed_monthly.iterrows():
			rate_gap = max(cycle_month_rate - float(month_row["Rebate_Mensual_Pct"]), 0.0)
			recoverable_value = float(month_row["Compra_Aplicable"]) * rate_gap
			if month_row["Cartera_Riesgo"]:
				blocked_pool += recoverable_value
			else:
				recoverable_pool += recoverable_value

	return {
		"Compra_Aplicable_Acumulada": applicable_purchase,
		"Meta_Acumulada_E1": target_elapsed_e1,
		"Meta_Acumulada_E2": target_elapsed_e2,
		"Meta_Ciclo_E1": target_total_e1,
		"Meta_Ciclo_E2": target_total_e2,
		"Cumplimiento_Acumulado_E1": safe_divide(applicable_purchase, target_elapsed_e1),
		"Cumplimiento_Acumulado_E2": safe_divide(applicable_purchase, target_elapsed_e2),
		"Escala_Ciclo": cycle_scale,
		"Faltante_E1_Actual": max(target_elapsed_e1 - applicable_purchase, 0.0),
		"Faltante_E2_Actual": max(target_elapsed_e2 - applicable_purchase, 0.0),
		"Faltante_E1_Ciclo": max(target_total_e1 - applicable_purchase, 0.0),
		"Faltante_E2_Ciclo": max(target_total_e2 - applicable_purchase, 0.0),
		"Recomposicion_9M_Proyectada": recoverable_pool * config["cycle_recomposition_factor"],
		"Recomposicion_9M_Bloqueada": blocked_pool,
	}


def generate_excel_report(executive_df: pd.DataFrame, monthly_df: pd.DataFrame, quarterly_df: pd.DataFrame, invoices_df: pd.DataFrame) -> io.BytesIO:
	sheets_to_write = [
		("Resumen_Ejecutivo", executive_df if not executive_df.empty else pd.DataFrame([{"Mensaje": "Sin información disponible"}])),
		("Mes_a_Mes", monthly_df if not monthly_df.empty else pd.DataFrame([{"Mensaje": "Sin información disponible"}])),
		("Trimestres", quarterly_df if not quarterly_df.empty else pd.DataFrame([{"Mensaje": "Sin información disponible"}])),
		("Facturas_Pintuco", invoices_df if not invoices_df.empty else pd.DataFrame([{"Mensaje": "Sin información disponible"}])),
	]
	return generate_excel_report_from_sheets(sheets_to_write)


def generate_excel_report_from_sheets(sheets_to_write: list[tuple[str, pd.DataFrame]]) -> io.BytesIO:
	output = io.BytesIO()
	workbook = Workbook()

	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="0B3C5D", end_color="0B3C5D", fill_type="solid")
	center_alignment = Alignment(horizontal="center", vertical="center")
	currency_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
	percent_format = "0.0%"

	for index, (sheet_name, dataframe) in enumerate(sheets_to_write):
		worksheet = workbook.active if index == 0 else workbook.create_sheet(title=sheet_name)
		worksheet.title = sheet_name
		worksheet.freeze_panes = "A2"

		prepared_df = dataframe.copy()
		for column in prepared_df.columns:
			if pd.api.types.is_datetime64_any_dtype(prepared_df[column]):
				prepared_df[column] = prepared_df[column].dt.strftime("%Y-%m-%d %H:%M:%S")

		rows = dataframe_to_rows(prepared_df, index=False, header=True)
		for row_index, row in enumerate(rows, start=1):
			for column_index, value in enumerate(row, start=1):
				cell = worksheet.cell(row=row_index, column=column_index, value=value)
				if row_index == 1:
					cell.font = header_font
					cell.fill = header_fill
					cell.alignment = center_alignment
				elif isinstance(value, (int, float)):
					column_name = prepared_df.columns[column_index - 1]
					if any(token in column_name for token in ["Compra", "Meta", "Faltante", "Rebate", "Exclusion", "Presupuesto", "Bono", "Pendiente", "Recomposicion"]):
						cell.number_format = currency_format
					if any(token in column_name for token in ["Cumplimiento", "Avance", "Pct"]):
						cell.number_format = percent_format

		for column_cells in worksheet.columns:
			max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
			worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

	workbook.save(output)
	output.seek(0)
	return output


def get_current_period_row(period_df: pd.DataFrame, snapshot_date: date) -> pd.Series:
	if period_df.empty:
		return pd.Series(dtype="object")

	current = period_df[
		(period_df["Inicio"].dt.date <= snapshot_date)
		& (period_df["Fin"].dt.date >= snapshot_date)
	]
	if not current.empty:
		return current.iloc[0]

	past = period_df[period_df["Fin"].dt.date < snapshot_date]
	if not past.empty:
		return past.iloc[-1]

	return period_df.iloc[0]


def render_provider_sync_panel(provider_config: dict[str, Any]) -> None:
	sync_col, sync_info_col = st.columns([1, 2])
	with sync_col:
		if st.button(f"🔄 Sincronizar facturas de {provider_config['label']}", type="primary", width="stretch", key=f"sync_{provider_config['key']}"):
			run_provider_sync(provider_config)
			st.cache_data.clear()
			st.rerun()
	with sync_info_col:
		last_sync_key = get_last_sync_key(provider_config)
		if last_sync_key in st.session_state:
			st.success(f"Última foto guardada: {st.session_state[last_sync_key]}")
		else:
			st.info(f"Aún no hay una foto guardada para {provider_config['label']}. Usa la sincronización para crear la primera foto del ciclo.")

	stats_key = get_last_sync_stats_key(provider_config)
	if stats_key in st.session_state:
		sync_stats = st.session_state[stats_key]
		st.caption(
			" | ".join(
				[
					f"Correos encontrados: {sync_stats.get('emails_found', 0)}",
					f"Correos procesados: {sync_stats.get('emails_processed', 0)}",
					f"Adjuntos revisados: {sync_stats.get('attachments_scanned', 0)}",
					f"XML leídos: {sync_stats.get('xml_files_scanned', 0)}",
					f"Facturas detectadas: {sync_stats.get('invoice_rows_detected', 0)}",
				]
			)
		)


def build_abracol_tracking_table(df: pd.DataFrame, budget_df: pd.DataFrame, snapshot_date: date) -> pd.DataFrame:
	rows = []
	working_df = df.copy()

	for _, budget_row in budget_df.iterrows():
		start = pd.Timestamp(budget_row["Inicio"])
		end = pd.Timestamp(budget_row["Fin"])
		period_df = working_df[
			(working_df["Fecha_Factura"].dt.date >= start.date())
			& (working_df["Fecha_Factura"].dt.date <= end.date())
		].copy()

		actual_sales = float(period_df["Valor_Neto"].sum()) if not period_df.empty else 0.0
		pending_value = float(period_df.loc[period_df["Estado_Pago"] == "Pendiente", "Valor_Neto"].sum()) if not period_df.empty else 0.0
		rebate_actual = actual_sales * ABRACOL_REBATE_RATE
		iva_actual = rebate_actual * VAT_RATE
		total_actual = rebate_actual + iva_actual
		budget_rebate = float(budget_row["Meta_2026"]) * ABRACOL_REBATE_RATE

		if snapshot_date < start.date():
			status = "Futuro"
		elif actual_sales >= float(budget_row["Meta_2026"]):
			status = "Cumplido"
		elif snapshot_date <= end.date():
			status = "Abierto"
		else:
			status = "No cumplido"

		rows.append(
			{
				"Periodo": budget_row["Periodo"],
				"Inicio": start,
				"Fin": end,
				"Ventas_2025": float(budget_row["Ventas_2025"]),
				"Meta_2026": float(budget_row["Meta_2026"]),
				"Ventas_Actuales_2026": actual_sales,
				"Cumplimiento_Meta": safe_divide(actual_sales, float(budget_row["Meta_2026"])),
				"Crecimiento_vs_2025": safe_divide(actual_sales, float(budget_row["Ventas_2025"])) - 1 if float(budget_row["Ventas_2025"]) else 0.0,
				"Faltante_Meta": max(float(budget_row["Meta_2026"]) - actual_sales, 0.0),
				"Rebate_Presupuestado": budget_rebate,
				"Rebate_Actual": rebate_actual,
				"IVA_Actual": iva_actual,
				"Total_Actual": total_actual,
				"Pendiente_Cartera": pending_value,
				"Rebate_Bloqueado": pending_value * ABRACOL_REBATE_RATE,
				"Estado_Periodo": status,
				"Facturas": int(period_df["Numero_Factura"].nunique()) if not period_df.empty else 0,
			}
		)

	return pd.DataFrame(rows)


def resolve_goya_tier(actual_sales: float, budget_row: pd.Series) -> tuple[str, float, float, str, float]:
	if actual_sales >= float(budget_row["Meta_50"]):
		return "50%", 0.50, 0.04, "Meta máxima alcanzada", 0.0
	if actual_sales >= float(budget_row["Meta_40"]):
		return "40%", 0.40, 0.035, "50%", max(float(budget_row["Meta_50"]) - actual_sales, 0.0)
	if actual_sales >= float(budget_row["Meta_30"]):
		return "30%", 0.30, 0.025, "40%", max(float(budget_row["Meta_40"]) - actual_sales, 0.0)
	if actual_sales >= float(budget_row["Meta_20"]):
		return "20%", 0.20, 0.02, "30%", max(float(budget_row["Meta_30"]) - actual_sales, 0.0)
	return "Sin escala", 0.0, 0.0, "20%", max(float(budget_row["Meta_20"]) - actual_sales, 0.0)


def build_goya_tracking_table(df: pd.DataFrame, budget_df: pd.DataFrame, snapshot_date: date) -> pd.DataFrame:
	rows = []
	working_df = df.copy()

	for _, budget_row in budget_df.iterrows():
		start = pd.Timestamp(budget_row["Inicio"])
		end = pd.Timestamp(budget_row["Fin"])
		period_df = working_df[
			(working_df["Fecha_Factura"].dt.date >= start.date())
			& (working_df["Fecha_Factura"].dt.date <= end.date())
		].copy()

		actual_sales = float(period_df["Valor_Neto"].sum()) if not period_df.empty else 0.0
		pending_value = float(period_df.loc[period_df["Estado_Pago"] == "Pendiente", "Valor_Neto"].sum()) if not period_df.empty else 0.0
		tier_label, growth_target, rebate_rate, next_target_label, missing_to_next = resolve_goya_tier(actual_sales, budget_row)
		rebate_actual = actual_sales * rebate_rate

		if snapshot_date < start.date():
			status = "Futuro"
		elif tier_label != "Sin escala":
			status = "Cumplido"
		elif snapshot_date <= end.date():
			status = "Abierto"
		else:
			status = "No cumplido"

		rows.append(
			{
				"Periodo": budget_row["Periodo"],
				"Inicio": start,
				"Fin": end,
				"Ventas_2024": float(budget_row["Ventas_2024"]),
				"Base_2025": float(budget_row["Base_2025"]),
				"Meta_20": float(budget_row["Meta_20"]),
				"Meta_30": float(budget_row["Meta_30"]),
				"Meta_40": float(budget_row["Meta_40"]),
				"Meta_50": float(budget_row["Meta_50"]),
				"Ventas_Actuales_2026": actual_sales,
				"Crecimiento_Actual": safe_divide(actual_sales, float(budget_row["Base_2025"])) - 1 if float(budget_row["Base_2025"]) else 0.0,
				"Escala_Lograda": tier_label,
				"Meta_Crecimiento_Lograda": growth_target,
				"Rebate_Pct": rebate_rate,
				"Rebate_Ganado": rebate_actual,
				"Pendiente_Cartera": pending_value,
				"Rebate_Bloqueado": pending_value * rebate_rate,
				"Siguiente_Meta": next_target_label,
				"Faltante_Siguiente_Meta": missing_to_next,
				"Estado_Periodo": status,
				"Facturas": int(period_df["Numero_Factura"].nunique()) if not period_df.empty else 0,
			}
		)

	return pd.DataFrame(rows)


def render_provider_invoice_tab(provider_config: dict[str, Any], provider_df: pd.DataFrame, snapshot_date: date, show_exclusion_columns: bool = False) -> None:
	st.subheader("Base factura por factura")
	cycle_start = provider_config["cycle_start"]
	filter_col1, filter_col2, filter_col3, filter_col4, filter_col5 = st.columns([1, 1, 1, 1.1, 1.4])
	with filter_col1:
		filter_start = st.date_input("Desde", value=cycle_start, min_value=cycle_start, max_value=snapshot_date, key=f"{provider_config['key']}_from")
	with filter_col2:
		filter_end = st.date_input("Hasta", value=snapshot_date, min_value=cycle_start, max_value=snapshot_date, key=f"{provider_config['key']}_to")
	with filter_col3:
		estado_options = ["Pendiente", "Pagada"]
		estado_filter = st.multiselect("Estado de pago", options=estado_options, default=estado_options, key=f"{provider_config['key']}_state")
	with filter_col4:
		visibility_filter = st.selectbox("Visibilidad", options=["Activas", "Excluidas", "Todas"], index=0, key=f"{provider_config['key']}_visibility")
	with filter_col5:
		search_term = st.text_input("Buscar factura o correo", placeholder="Factura, remitente, asunto o adjunto", key=f"{provider_config['key']}_search")

	if filter_end < filter_start:
		st.error("La fecha final no puede ser menor que la fecha inicial.")
		st.stop()

	provider_df = provider_df.copy()
	provider_df["Fecha_Exclusion"] = normalize_datetime_series(provider_df.get("Fecha_Exclusion", pd.Series(index=provider_df.index, dtype=object)))
	provider_df["Excluir_De_Calculos"] = provider_df.get("Excluir_De_Calculos", pd.Series(False, index=provider_df.index)).astype(bool)
	active_count = int((~provider_df["Excluir_De_Calculos"]).sum())
	excluded_count = int(provider_df["Excluir_De_Calculos"].sum())
	st.caption(f"Facturas activas: {active_count:,} · Facturas excluidas de cálculos: {excluded_count:,}")

	filtered_df = provider_df[
		(provider_df["Fecha_Factura"].dt.date >= filter_start) & (provider_df["Fecha_Factura"].dt.date <= filter_end)
	].copy()

	if visibility_filter == "Activas":
		filtered_df = filtered_df[~filtered_df["Excluir_De_Calculos"]].copy()
	elif visibility_filter == "Excluidas":
		filtered_df = filtered_df[filtered_df["Excluir_De_Calculos"]].copy()

	if estado_filter:
		filtered_df = filtered_df[filtered_df["Estado_Pago"].isin(estado_filter)].copy()

	if search_term.strip():
		search_value = search_term.strip().upper()
		search_columns = ["Numero_Factura", "Proveedor_Correo", "Remitente_Correo", "Asunto_Correo", "Nombre_Adjunto", "Message_ID"]
		search_mask = pd.Series(False, index=filtered_df.index)
		for column in search_columns:
			search_mask = search_mask | filtered_df[column].fillna("").astype(str).str.upper().str.contains(search_value, regex=False)
		filtered_df = filtered_df[search_mask].copy()

	if filtered_df.empty:
		st.warning("No hay facturas en el rango seleccionado.")
		return

	display_df = sort_invoice_dataframe(filtered_df, by=["Fecha_Factura", "Fecha_Recepcion_Correo"], ascending=[False, False])
	column_config = {
		"Fecha_Factura": st.column_config.DateColumn("Fecha factura", format="YYYY-MM-DD"),
		"Valor_Neto": st.column_config.NumberColumn("Compra neta", format="$ %,.0f"),
		"Fecha_Recepcion_Correo": st.column_config.DatetimeColumn("Fecha correo", format="YYYY-MM-DD HH:mm"),
		"Estado_Pago": st.column_config.TextColumn("Estado de pago"),
		"Excluir_De_Calculos": st.column_config.CheckboxColumn("Excluir"),
		"Fecha_Exclusion": st.column_config.DatetimeColumn("Fecha exclusión", format="YYYY-MM-DD HH:mm"),
	}
	if show_exclusion_columns:
		column_config["Compra_Excluida_12"] = st.column_config.NumberColumn("12% excluido", format="$ %,.0f")
		column_config["Compra_Aplicable_Rebate"] = st.column_config.NumberColumn("88% aplicable", format="$ %,.0f")

	visible_columns = [column for column in PROVIDER_EXCLUSION_DISPLAY_COLUMNS if column in display_df.columns]
	if show_exclusion_columns:
		visible_columns.extend([column for column in ["Compra_Excluida_12", "Compra_Aplicable_Rebate"] if column in display_df.columns])

	st.dataframe(display_df[visible_columns], width="stretch", hide_index=True, column_config=column_config)

	st.markdown("---")
	st.markdown("**Control de exclusiones**")
	st.caption("Marca aquí las facturas promocionales o no operativas para que dejen de sumar en rebates y en el resto de vistas de la app. La exclusión queda guardada en Google Sheets y se puede revertir.")

	active_candidates = sort_invoice_dataframe(provider_df[~provider_df["Excluir_De_Calculos"]].copy(), by=["Fecha_Factura", "Numero_Factura"], ascending=[False, False])
	excluded_candidates = sort_invoice_dataframe(provider_df[provider_df["Excluir_De_Calculos"]].copy(), by=["Fecha_Exclusion", "Fecha_Factura"], ascending=[False, False])

	manage_col1, manage_col2 = st.columns(2)
	with manage_col1:
		st.markdown("Excluir factura")
		if active_candidates.empty:
			st.info("No hay facturas activas disponibles para excluir.")
		else:
			exclude_options = {
				f"{row['Numero_Factura']} | {row['Fecha_Factura'].date() if pd.notna(row['Fecha_Factura']) else 'Sin fecha'} | {format_currency(float(row['Valor_Neto']))}": row["Invoice_Key"]
				for _, row in active_candidates.iterrows()
			}
			selected_exclude_label = st.selectbox("Factura a excluir", list(exclude_options.keys()), key=f"{provider_config['key']}_exclude_invoice")
			exclude_reason = st.text_input("Motivo", value="Factura promocional / no tener en cuenta", key=f"{provider_config['key']}_exclude_reason")
			selected_exclude_key = exclude_options[selected_exclude_label]
			selected_exclude_row = active_candidates[active_candidates["Invoice_Key"] == selected_exclude_key].iloc[0]
			if st.button("Guardar exclusión", type="primary", width="stretch", key=f"{provider_config['key']}_save_exclusion"):
				gs_client = connect_to_google_sheets()
				if not gs_client:
					st.error("No fue posible conectar con Google Sheets para guardar la exclusión.")
				elif register_invoice_exclusion(
					gs_client,
					invoice_key=selected_exclude_row["Invoice_Key"],
					proveedor_norm=normalize_supplier_key(provider_config["provider_names_erp"][0]),
					num_factura=selected_exclude_row["Numero_Factura"],
					reason=exclude_reason,
					source=f"rebate_{provider_config['key']}",
				):
					load_invoice_exclusion_registry.clear()
					load_provider_data_from_gsheet.clear()
					st.session_state.pop("treasury_payload", None)
					st.success("Factura marcada para no tener en cuenta. Desde ahora deja de sumar en los cálculos de la app.")
					st.rerun()
				else:
					st.error("No se pudo guardar la exclusión.")

	with manage_col2:
		st.markdown("Revertir exclusión")
		if excluded_candidates.empty:
			st.info("No hay facturas excluidas para este proveedor.")
		else:
			restore_options = {
				f"{row['Numero_Factura']} | {row['Fecha_Factura'].date() if pd.notna(row['Fecha_Factura']) else 'Sin fecha'} | {row['Motivo_Exclusion'] or 'Sin motivo'}": row["Exclusion_ID"]
				for _, row in excluded_candidates.iterrows()
			}
			selected_restore_label = st.selectbox("Factura a reincluir", list(restore_options.keys()), key=f"{provider_config['key']}_restore_invoice")
			selected_restore_id = restore_options[selected_restore_label]
			if st.button("Quitar exclusión", width="stretch", key=f"{provider_config['key']}_restore_exclusion"):
				gs_client = connect_to_google_sheets()
				if not gs_client:
					st.error("No fue posible conectar con Google Sheets para revertir la exclusión.")
				elif deactivate_invoice_exclusion(gs_client, selected_restore_id):
					load_invoice_exclusion_registry.clear()
					load_provider_data_from_gsheet.clear()
					st.session_state.pop("treasury_payload", None)
					st.success("La factura volvió a quedar activa en todos los cálculos.")
					st.rerun()
				else:
					st.error("No se pudo revertir la exclusión.")


def render_provider_diagnostics_tab(provider_label: str, provider_df: pd.DataFrame, alerts: list[str], export_sheets: list[tuple[str, pd.DataFrame]], file_name: str) -> None:
	st.subheader("Diagnóstico operativo y descarga")
	unique_senders = provider_df["Remitente_Correo"].replace("", pd.NA).dropna().nunique()
	unique_subjects = provider_df["Asunto_Correo"].replace("", pd.NA).dropna().nunique()
	duplicate_invoices = provider_df["Numero_Factura"].duplicated().sum()
	trazabilidad_completa = (
		provider_df["Fecha_Recepcion_Correo"].notna()
		& provider_df["Remitente_Correo"].fillna("").astype(str).str.strip().ne("")
		& provider_df["Message_ID"].fillna("").astype(str).str.strip().ne("")
	).mean()

	render_kpi_grid(
		[
			kpi_card_html("Cobertura de trazabilidad", format_percent(float(trazabilidad_completa), 0), "Correo, fecha y Message-ID", "navy"),
			kpi_card_html("Remitentes identificados", f"{unique_senders:,}", "Origen de correo detectado", "green"),
			kpi_card_html("Asuntos distintos", f"{unique_subjects:,}", "Control de buzón", "gold"),
			kpi_card_html("Duplicados visibles", f"{duplicate_invoices:,}", "Facturas repetidas en la base", "red"),
		]
	)

	st.markdown(f"Alertas del seguimiento de {provider_label}")
	for alert in alerts:
		if "no se detectan alertas" in alert.lower():
			st.success(alert)
		else:
			st.warning(alert)

	excel_data = generate_excel_report_from_sheets(export_sheets)
	st.download_button(
		label="⬇️ Descargar consolidado ejecutivo del rebate",
		data=excel_data,
		file_name=file_name,
		mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		width="stretch",
	)


def render_abracol_dashboard(provider_config: dict[str, Any]) -> None:
	st.markdown(
		f"""
		<div class="pintuco-banner">
			<h1>{provider_config['title']}</h1>
			<p>{provider_config['cycle_name']}. Seguimiento bimestral del rebate Gold de Abracol, con cruce de correo y ERP para ver ventas, cumplimiento, rebate y total con IVA.</p>
		</div>
		""",
		unsafe_allow_html=True,
	)

	st.markdown(
		"""
		<div class="info-card">
			<strong>Reglas comerciales activas</strong><br>
			1. Abracol liquida <strong>6%</strong> por bimestre sobre venta neta facturada.<br>
			2. La <strong>nota se reconoce antes de IVA</strong>; el tablero muestra rebate bruto, IVA estimado y total de la nota.<br>
			3. La cuota comercial 2026 se compara contra la meta bimestral definida en el cuadro cargado por Ferreinox.<br>
			4. El cruce de cartera separa el valor ya facturado del valor aún pendiente de pago para lectura financiera.
		</div>
		""",
		unsafe_allow_html=True,
	)

	render_provider_sync_panel(provider_config)
	exclusion_df = load_invoice_exclusion_registry()
	abracol_df = load_provider_data_from_gsheet(provider_config["worksheet_name"], provider_config["cycle_start"], provider_config["excluded_purchase_percent"])
	abracol_df = apply_provider_invoice_exclusions(abracol_df, provider_config, exclusion_df)
	if abracol_df.empty:
		st.warning("No hay datos de Abracol para el ciclo vigente. Ejecuta la sincronización inicial.")
		return
	abracol_active_df, abracol_excluded_df = split_provider_dataset(abracol_df)
	if abracol_active_df.empty and not abracol_excluded_df.empty:
		st.warning("Todas las facturas visibles de Abracol quedaron excluidas de cálculos. Revisa la pestaña de facturas para reactivar las que sí deben contar.")
		return

	snapshot_date = datetime.now(COLOMBIA_TZ).date()
	period_df = build_abracol_tracking_table(abracol_active_df, provider_config["budget_df"], snapshot_date)
	tracking_alerts = build_tracking_alerts(abracol_active_df)
	current_period = get_current_period_row(period_df, snapshot_date)
	ventas_acumuladas = float(period_df["Ventas_Actuales_2026"].sum())
	meta_total = float(period_df["Meta_2026"].sum())
	rebate_total = float(period_df["Rebate_Actual"].sum())
	rebate_blocked = float(period_df["Rebate_Bloqueado"].sum())

	render_kpi_grid(
		[
			kpi_card_html("Ventas 2026 facturadas", format_currency(ventas_acumuladas), f"Cumplimiento anual: {format_percent(safe_divide(ventas_acumuladas, meta_total))}", "navy"),
			kpi_card_html("Bimestre actual", str(current_period["Periodo"]), f"Cumplimiento meta: {format_percent(float(current_period['Cumplimiento_Meta']))}", get_scale_tone(str(current_period["Estado_Periodo"]))),
			kpi_card_html("Rebate bruto estimado", format_currency(rebate_total), f"Abracol 6.0% antes de IVA", "green"),
			kpi_card_html("Rebate bloqueado cartera", format_currency(rebate_blocked), f"Pendiente cartera: {format_currency(float(period_df['Pendiente_Cartera'].sum()))}", "gold"),
		]
	)

	render_kpi_grid(
		[
			kpi_card_html("Meta bimestre actual", format_currency(float(current_period["Meta_2026"])), f"Ventas 2025: {format_currency(float(current_period['Ventas_2025']))}", "navy"),
			kpi_card_html("Ventas actuales", format_currency(float(current_period["Ventas_Actuales_2026"])), f"Crecimiento vs 2025: {format_percent(float(current_period['Crecimiento_vs_2025']))}", "green"),
			kpi_card_html("IVA estimado nota", format_currency(float(current_period["IVA_Actual"])), f"Total con IVA: {format_currency(float(current_period['Total_Actual']))}", "gold"),
			kpi_card_html("Estado del periodo", str(current_period["Estado_Periodo"]), f"Faltante meta: {format_currency(float(current_period['Faltante_Meta']))}", get_scale_tone(str(current_period["Estado_Periodo"]))),
		]
	)

	st.markdown(
		"""
		<div class="note-card">
			<strong>Lectura ejecutiva</strong><br>
			Abracol se presenta contra su cuota bimestral. El tablero muestra cuánto de la meta ya está facturado, cuánto rebate bruto genera ese volumen y cuánto está todavía bloqueado por cartera pendiente.
		</div>
		""",
		unsafe_allow_html=True,
	)

	overview_tab, period_tab, invoices_tab, diagnostics_tab = st.tabs(["📊 Dirección", "🗓️ Bimestres", "📑 Facturas y fuente", "🛰️ Diagnóstico"])

	with overview_tab:
		st.subheader("Ritmo bimestral")
		chart_df = period_df.set_index("Periodo")[["Ventas_Actuales_2026", "Meta_2026"]]
		st.line_chart(chart_df, width="stretch")
		overview_col1, overview_col2 = st.columns([1.2, 1])
		with overview_col1:
			st.markdown(
				f"""
				<div class="section-card">
					<strong>{current_period['Periodo']}</strong><br><br>
					Meta 2026: <strong>{format_currency(float(current_period['Meta_2026']))}</strong><br>
					Ventas actuales: <strong>{format_currency(float(current_period['Ventas_Actuales_2026']))}</strong><br>
					Ventas 2025: <strong>{format_currency(float(current_period['Ventas_2025']))}</strong><br>
					Faltante meta: <strong>{format_currency(float(current_period['Faltante_Meta']))}</strong><br>
					Estado: {pill_html(str(current_period['Estado_Periodo']), get_scale_tone(str(current_period['Estado_Periodo'])))}
				</div>
				""",
				unsafe_allow_html=True,
			)
		with overview_col2:
			st.markdown(
				f"""
				<div class="section-card">
					<strong>Rebate del periodo</strong><br><br>
					Rebate bruto: <strong>{format_currency(float(current_period['Rebate_Actual']))}</strong><br>
					IVA estimado: <strong>{format_currency(float(current_period['IVA_Actual']))}</strong><br>
					Total con IVA: <strong>{format_currency(float(current_period['Total_Actual']))}</strong><br>
					Bloqueado por cartera: <strong>{format_currency(float(current_period['Rebate_Bloqueado']))}</strong>
				</div>
				""",
				unsafe_allow_html=True,
			)

	with period_tab:
		st.subheader("Seguimiento bimestral")
		st.dataframe(
			period_df[[
				"Periodo",
				"Ventas_2025",
				"Meta_2026",
				"Ventas_Actuales_2026",
				"Cumplimiento_Meta",
				"Crecimiento_vs_2025",
				"Faltante_Meta",
				"Rebate_Actual",
				"IVA_Actual",
				"Total_Actual",
				"Pendiente_Cartera",
				"Estado_Periodo",
			]],
			width="stretch",
			hide_index=True,
			column_config={
				"Ventas_2025": st.column_config.NumberColumn("Ventas 2025", format="$ %,.0f"),
				"Meta_2026": st.column_config.NumberColumn("Meta 2026", format="$ %,.0f"),
				"Ventas_Actuales_2026": st.column_config.NumberColumn("Ventas actuales", format="$ %,.0f"),
				"Cumplimiento_Meta": st.column_config.ProgressColumn("Cumplimiento", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Crecimiento_vs_2025": st.column_config.ProgressColumn("Crec. vs 2025", min_value=-0.2, max_value=0.5, format="%.0f%%"),
				"Faltante_Meta": st.column_config.NumberColumn("Faltante meta", format="$ %,.0f"),
				"Rebate_Actual": st.column_config.NumberColumn("Rebate", format="$ %,.0f"),
				"IVA_Actual": st.column_config.NumberColumn("IVA", format="$ %,.0f"),
				"Total_Actual": st.column_config.NumberColumn("Total", format="$ %,.0f"),
				"Pendiente_Cartera": st.column_config.NumberColumn("Pendiente cartera", format="$ %,.0f"),
			},
		)
		st.bar_chart(period_df.set_index("Periodo")[["Rebate_Actual", "IVA_Actual"]], width="stretch")

	with invoices_tab:
		render_provider_invoice_tab(provider_config, abracol_df, snapshot_date)

	with diagnostics_tab:
		executive_export_df = pd.DataFrame([
			{
				"Foto": datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M:%S"),
				"Ventas_2026_Facturadas": ventas_acumuladas,
				"Meta_2026_Anual": meta_total,
				"Cumplimiento_Anual": safe_divide(ventas_acumuladas, meta_total),
				"Rebate_Bruto_Estimado": rebate_total,
				"Rebate_Bloqueado_Cartera": rebate_blocked,
			}
		])
		render_provider_diagnostics_tab(
			provider_config["label"],
			abracol_df,
			tracking_alerts,
			[("Resumen_Ejecutivo", executive_export_df), ("Bimestres", period_df), ("Facturas_Abracol", abracol_df)],
			f"Rebate_Abracol_Ejecutivo_{snapshot_date}.xlsx",
		)


def render_goya_dashboard(provider_config: dict[str, Any]) -> None:
	st.markdown(
		f"""
		<div class="pintuco-banner">
			<h1>{provider_config['title']}</h1>
			<p>{provider_config['cycle_name']}. Seguimiento semestral del crecimiento Goya contra base 2025, con liquidación por escalas y control documental sobre correo y ERP.</p>
		</div>
		""",
		unsafe_allow_html=True,
	)

	st.markdown(
		f"""
		<div class="info-card">
			<strong>Reglas comerciales activas</strong><br>
			1. Goya liquida rebate por <strong>semestre</strong>: enero-junio y julio-diciembre.<br>
			2. La base de comparación es la venta 2025 del semestre; las escalas son <strong>20%, 30%, 40% y 50%</strong> de crecimiento.<br>
			3. Porcentaje de rebate aplicado sobre la venta vigente: <strong>2.0%, 2.5%, 3.5% y 4.0%</strong>.<br>
			4. El cálculo usa compras reales facturadas del periodo. Las condiciones estratégicas por mezcla de líneas se muestran como nota comercial, no como bloqueo automático.<br>
			5. <strong>Nota documental:</strong> {provider_config['note']}
		</div>
		""",
		unsafe_allow_html=True,
	)

	render_provider_sync_panel(provider_config)
	exclusion_df = load_invoice_exclusion_registry()
	goya_df = load_provider_data_from_gsheet(provider_config["worksheet_name"], provider_config["cycle_start"], provider_config["excluded_purchase_percent"])
	goya_df = apply_provider_invoice_exclusions(goya_df, provider_config, exclusion_df)
	if goya_df.empty:
		st.warning("No hay datos de Goya para el ciclo vigente. Ejecuta la sincronización inicial.")
		return
	goya_active_df, goya_excluded_df = split_provider_dataset(goya_df)
	if goya_active_df.empty and not goya_excluded_df.empty:
		st.warning("Todas las facturas visibles de Goya quedaron excluidas de cálculos. Revisa la pestaña de facturas para reactivar las que sí deben contar.")
		return

	snapshot_date = datetime.now(COLOMBIA_TZ).date()
	period_df = build_goya_tracking_table(goya_active_df, provider_config["budget_df"], snapshot_date)
	tracking_alerts = build_tracking_alerts(goya_active_df)
	current_period = get_current_period_row(period_df, snapshot_date)
	ventas_acumuladas = float(period_df["Ventas_Actuales_2026"].sum())
	base_total = float(period_df["Base_2025"].sum())
	rebate_total = float(period_df["Rebate_Ganado"].sum())
	rebate_blocked = float(period_df["Rebate_Bloqueado"].sum())

	render_kpi_grid(
		[
			kpi_card_html("Ventas 2026 facturadas", format_currency(ventas_acumuladas), f"Crecimiento consolidado: {format_percent(safe_divide(ventas_acumuladas, base_total) - 1 if base_total else 0.0)}", "navy"),
			kpi_card_html("Semestre actual", str(current_period["Periodo"]), f"Escala lograda: {current_period['Escala_Lograda']}", get_scale_tone(str(current_period["Escala_Lograda"]))),
			kpi_card_html("Rebate ganado", format_currency(rebate_total), f"Pagadero en producto según acuerdo", "green"),
			kpi_card_html("Rebate bloqueado cartera", format_currency(rebate_blocked), f"Pendiente cartera: {format_currency(float(period_df['Pendiente_Cartera'].sum()))}", "gold"),
		]
	)

	render_kpi_grid(
		[
			kpi_card_html("Base 2025 semestre", format_currency(float(current_period["Base_2025"])), f"Referencia 2024: {format_currency(float(current_period['Ventas_2024']))}", "navy"),
			kpi_card_html("Ventas actuales semestre", format_currency(float(current_period["Ventas_Actuales_2026"])), f"Crecimiento actual: {format_percent(float(current_period['Crecimiento_Actual']))}", "green"),
			kpi_card_html("Tasa rebate aplicada", format_percent(float(current_period["Rebate_Pct"]), 1), f"Siguiente meta: {current_period['Siguiente_Meta']}", "gold"),
			kpi_card_html("Faltante siguiente meta", format_currency(float(current_period["Faltante_Siguiente_Meta"])), f"Estado: {current_period['Estado_Periodo']}", get_scale_tone(str(current_period["Estado_Periodo"]))),
		]
	)

	st.markdown(
		"""
		<div class="note-card">
			<strong>Lectura ejecutiva</strong><br>
			El tablero compara cada semestre 2026 contra su base 2025, determina la escala lograda y proyecta el rebate sobre la venta facturada. Las condiciones estratégicas por mezcla de líneas y crecimiento en unidades quedan visibles como criterio comercial, pero no se bloquean automáticamente porque hoy no hay detalle SKU en la fuente documental.
		</div>
		""",
		unsafe_allow_html=True,
	)

	overview_tab, period_tab, invoices_tab, diagnostics_tab = st.tabs(["📊 Dirección", "🗓️ Semestres", "📑 Facturas y fuente", "🛰️ Diagnóstico"])

	with overview_tab:
		st.subheader("Ritmo semestral")
		chart_df = period_df.set_index("Periodo")[["Ventas_Actuales_2026", "Meta_20", "Meta_30", "Meta_40", "Meta_50"]]
		st.line_chart(chart_df, width="stretch")
		overview_col1, overview_col2 = st.columns([1.15, 1])
		with overview_col1:
			st.markdown(
				f"""
				<div class="section-card">
					<strong>{current_period['Periodo']}</strong><br><br>
					Base 2025: <strong>{format_currency(float(current_period['Base_2025']))}</strong><br>
					Ventas 2026: <strong>{format_currency(float(current_period['Ventas_Actuales_2026']))}</strong><br>
					Meta 20%: <strong>{format_currency(float(current_period['Meta_20']))}</strong><br>
					Meta 50%: <strong>{format_currency(float(current_period['Meta_50']))}</strong><br>
					Escala lograda: {pill_html(str(current_period['Escala_Lograda']), get_scale_tone(str(current_period['Escala_Lograda'])))}
				</div>
				""",
				unsafe_allow_html=True,
			)
		with overview_col2:
			st.markdown(
				f"""
				<div class="section-card">
					<strong>Condiciones comerciales visibles</strong><br><br>
					Rebate ganado: <strong>{format_currency(float(current_period['Rebate_Ganado']))}</strong><br>
					Bloqueado cartera: <strong>{format_currency(float(current_period['Rebate_Bloqueado']))}</strong><br>
					Siguiente meta: <strong>{current_period['Siguiente_Meta']}</strong><br>
					Faltante siguiente meta: <strong>{format_currency(float(current_period['Faltante_Siguiente_Meta']))}</strong>
				</div>
				""",
				unsafe_allow_html=True,
			)

	with period_tab:
		st.subheader("Seguimiento semestral")
		st.dataframe(
			period_df[[
				"Periodo",
				"Ventas_2024",
				"Base_2025",
				"Ventas_Actuales_2026",
				"Crecimiento_Actual",
				"Meta_20",
				"Meta_30",
				"Meta_40",
				"Meta_50",
				"Escala_Lograda",
				"Rebate_Pct",
				"Rebate_Ganado",
				"Pendiente_Cartera",
				"Estado_Periodo",
			]],
			width="stretch",
			hide_index=True,
			column_config={
				"Ventas_2024": st.column_config.NumberColumn("Ventas 2024", format="$ %,.0f"),
				"Base_2025": st.column_config.NumberColumn("Base 2025", format="$ %,.0f"),
				"Ventas_Actuales_2026": st.column_config.NumberColumn("Ventas 2026", format="$ %,.0f"),
				"Crecimiento_Actual": st.column_config.ProgressColumn("Crecimiento", min_value=-0.2, max_value=0.6, format="%.0f%%"),
				"Meta_20": st.column_config.NumberColumn("Meta 20%", format="$ %,.0f"),
				"Meta_30": st.column_config.NumberColumn("Meta 30%", format="$ %,.0f"),
				"Meta_40": st.column_config.NumberColumn("Meta 40%", format="$ %,.0f"),
				"Meta_50": st.column_config.NumberColumn("Meta 50%", format="$ %,.0f"),
				"Rebate_Pct": st.column_config.NumberColumn("% Rebate", format="%.1f%%"),
				"Rebate_Ganado": st.column_config.NumberColumn("Rebate ganado", format="$ %,.0f"),
				"Pendiente_Cartera": st.column_config.NumberColumn("Pendiente cartera", format="$ %,.0f"),
			},
		)

	with invoices_tab:
		render_provider_invoice_tab(provider_config, goya_df, snapshot_date)

	with diagnostics_tab:
		executive_export_df = pd.DataFrame([
			{
				"Foto": datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M:%S"),
				"Ventas_2026_Facturadas": ventas_acumuladas,
				"Base_2025_Acumulada": base_total,
				"Crecimiento_Acumulado": safe_divide(ventas_acumuladas, base_total) - 1 if base_total else 0.0,
				"Rebate_Ganado": rebate_total,
				"Rebate_Bloqueado_Cartera": rebate_blocked,
			}
		])
		render_provider_diagnostics_tab(
			provider_config["label"],
			goya_df,
			tracking_alerts,
			[("Resumen_Ejecutivo", executive_export_df), ("Semestres", period_df), ("Facturas_Goya", goya_df)],
			f"Rebate_Goya_Ejecutivo_{snapshot_date}.xlsx",
		)


def render_pintuco_dashboard(provider_config: dict[str, Any]) -> None:
	st.markdown(
		f"""
		<div class="pintuco-banner">
			<h1>{provider_config['title']}</h1>
			<p>{provider_config['cycle_name']}. Seguimiento comercial, financiero y operativo del presupuesto Ferreinox con lectura directa de facturas, estacionalidad, escalas, trimestre y recomposición.</p>
		</div>
		""",
		unsafe_allow_html=True,
	)

	st.markdown(
		"""
		<div class="info-card">
			<strong>Reglas comerciales activas</strong><br>
			1. La base del cálculo es la compra neta menos el <strong>12% excluido</strong>; el rebate corre sobre el <strong>88% aplicable</strong>.<br>
			2. Cada mes compara contra <strong>Escala 1</strong> y <strong>Escala 2</strong>; Escala 2 paga más que Escala 1.<br>
			3. El pago mensual y trimestral se calcula sobre <strong>toda la compra aplicable del periodo</strong>, sin techo.<br>
			4. Estacionalidad: si al cierre del <strong>tercer domingo</strong> del mes se alcanza al menos el <strong>90% de Escala 2</strong>, se gana un <strong>1% adicional</strong> sobre la compra aplicable del mes.<br>
			5. La recomposición separa el saldo <strong>elegible</strong> del saldo <strong>bloqueado por cartera</strong> para no mezclar comercial con cobranza.
		</div>
		""",
		unsafe_allow_html=True,
	)

	config = get_rebate_configuration()
	budget_df = config["budget_df"]
	snapshot_date = datetime.now(COLOMBIA_TZ).date()
	render_provider_sync_panel(provider_config)

	exclusion_df = load_invoice_exclusion_registry()
	pintuco_df = load_pintuco_data_from_gsheet()
	pintuco_df = apply_provider_invoice_exclusions(pintuco_df, provider_config, exclusion_df)
	if pintuco_df.empty:
		st.warning("No hay datos de Pintuco para el ciclo vigente. Ejecuta la sincronización inicial.")
		return
	pintuco_active_df, pintuco_excluded_df = split_provider_dataset(pintuco_df)
	if pintuco_active_df.empty and not pintuco_excluded_df.empty:
		st.warning("Todas las facturas visibles de Pintuco quedaron excluidas de cálculos. Revisa la pestaña de facturas para reactivar las que sí deben contar.")
		return

	monthly_df = build_monthly_rebate_table(pintuco_active_df, budget_df, config, snapshot_date)
	quarterly_df = build_quarterly_rebate_table(monthly_df, config)
	cycle_outlook = build_cycle_projection(monthly_df, budget_df, config, snapshot_date)
	summary_df = build_cycle_summary(monthly_df)
	tracking_alerts = build_tracking_alerts(pintuco_active_df)

	current_month_start = pd.Timestamp(snapshot_date.replace(day=1))
	current_month_df = monthly_df[monthly_df["Mes_Inicio"] == current_month_start]
	if current_month_df.empty:
		current_month_df = monthly_df.iloc[[0]]
	current_month = current_month_df.iloc[0]
	current_quarter = str(current_month["Trimestre"])
	current_quarter_df = quarterly_df[quarterly_df["Trimestre"] == current_quarter]
	if current_quarter_df.empty:
		current_quarter_df = quarterly_df.iloc[[0]]
	current_quarter_row = current_quarter_df.iloc[0]
	rebate_total_ganado = float(monthly_df["Rebate_Mensual_Ganado"].sum() + monthly_df["Bono_Estacionalidad"].sum() + quarterly_df["Rebate_Trimestral_Ganado"].sum())

	render_kpi_grid(
		[
			kpi_card_html("Compra aplicable acumulada", format_currency(cycle_outlook["Compra_Aplicable_Acumulada"]), f"Cumplimiento E2 acumulado: {format_percent(cycle_outlook['Cumplimiento_Acumulado_E2'])}", "navy"),
			kpi_card_html("Escala actual del ciclo", cycle_outlook["Escala_Ciclo"], f"Faltante acumulado a E2: {format_currency(cycle_outlook['Faltante_E2_Actual'])}", get_scale_tone(cycle_outlook["Escala_Ciclo"])),
			kpi_card_html("Rebate ganado a hoy", format_currency(rebate_total_ganado), "Mensual + trimestral + estacionalidad", "green"),
			kpi_card_html("Bolsa 9M proyectada", format_currency(cycle_outlook["Recomposicion_9M_Proyectada"]), f"Bloqueada por cartera: {format_currency(cycle_outlook['Recomposicion_9M_Bloqueada'])}", "gold"),
		]
	)

	render_kpi_grid(
		[
			kpi_card_html("Mes actual", format_currency(float(current_month["Compra_Aplicable"])), f"{current_month['Mes']} · escala {current_month['Escala_Lograda']}", get_scale_tone(str(current_month["Escala_Lograda"]))),
			kpi_card_html("Estacionalidad", str(current_month["Estado_Estacionalidad"]), f"Corte: {pd.Timestamp(current_month['Corte_Estacionalidad']).strftime('%Y-%m-%d')} · bono: {format_currency(float(current_month['Bono_Estacionalidad']))}", get_scale_tone(str(current_month["Estado_Estacionalidad"]))),
			kpi_card_html("Trimestre actual", str(current_quarter_row["Escala_Lograda"]), f"{current_quarter} · faltante E2: {format_currency(float(current_quarter_row['Faltante_E2']))}", get_scale_tone(str(current_quarter_row["Escala_Lograda"]))),
			kpi_card_html("Recomposición trimestral", format_currency(float(current_quarter_row["Recomposicion_Trimestral_Proyectada"])), f"Bloqueada cartera: {format_currency(float(current_quarter_row['Recomposicion_Cartera_Bloqueada']))}", "gold"),
		]
	)

	st.markdown(
		"""
		<div class="note-card">
			<strong>Lectura ejecutiva</strong><br>
			El tablero separa lo ya ganado del saldo recuperable. Cuando un mes o trimestre todavía tiene facturas pendientes, ese valor queda marcado como <strong>bloqueado por cartera</strong> y no se suma a la recomposición automática hasta que el riesgo operativo desaparezca.
		</div>
		""",
		unsafe_allow_html=True,
	)

	overview_tab, monthly_tab, quarter_tab, invoices_tab, diagnostics_tab = st.tabs(["📊 Dirección", "🗓️ Mes a mes", "📦 Trimestre y recomposición", "📑 Facturas y fuente", "🛰️ Diagnóstico"])

	with overview_tab:
		st.subheader("Ritmo del ciclo")
		chart_df = monthly_df.set_index("Mes")[["Compra_Aplicable", "Presupuesto_Escala_1", "Presupuesto_Escala_2"]]
		st.line_chart(chart_df, width="stretch")
		overview_col1, overview_col2 = st.columns([1.2, 1])
		with overview_col1:
			st.markdown(
				f"""
				<div class="section-card">
					<strong>Mes actual: {current_month['Mes']}</strong><br><br>
					Base aplicable: <strong>{format_currency(float(current_month['Compra_Aplicable']))}</strong><br>
					Presupuesto Escala 1: <strong>{format_currency(float(current_month['Presupuesto_Escala_1']))}</strong><br>
					Presupuesto Escala 2: <strong>{format_currency(float(current_month['Presupuesto_Escala_2']))}</strong><br>
					Faltante a E1: <strong>{format_currency(float(current_month['Faltante_E1']))}</strong><br>
					Faltante a E2: <strong>{format_currency(float(current_month['Faltante_E2']))}</strong><br>
					Escala lograda: {pill_html(str(current_month['Escala_Lograda']), get_scale_tone(str(current_month['Escala_Lograda'])))}
				</div>
				""",
				unsafe_allow_html=True,
			)
		with overview_col2:
			st.markdown(
				f"""
				<div class="section-card">
					<strong>{current_quarter}</strong><br><br>
					Compra aplicable: <strong>{format_currency(float(current_quarter_row['Compra_Aplicable']))}</strong><br>
					Rebate trimestral ganado: <strong>{format_currency(float(current_quarter_row['Rebate_Trimestral_Ganado']))}</strong><br>
					Recomposición trimestral elegible: <strong>{format_currency(float(current_quarter_row['Recomposicion_Trimestral_Proyectada']))}</strong><br>
					Recomposición bloqueada por cartera: <strong>{format_currency(float(current_quarter_row['Recomposicion_Cartera_Bloqueada']))}</strong><br>
					Escala lograda: {pill_html(str(current_quarter_row['Escala_Lograda']), get_scale_tone(str(current_quarter_row['Escala_Lograda'])))}
				</div>
				""",
				unsafe_allow_html=True,
			)

		st.subheader("Resumen ejecutivo del ciclo")
		st.dataframe(
			summary_df,
			width="stretch",
			hide_index=True,
			column_config={
				"Compra_Neta": st.column_config.NumberColumn("Compra neta", format="$ %,.0f"),
				"Exclusion_12": st.column_config.NumberColumn("12% excluido", format="$ %,.0f"),
				"Compra_Aplicable": st.column_config.NumberColumn("88% aplicable", format="$ %,.0f"),
				"Presupuesto_Escala_1": st.column_config.NumberColumn("Escala 1", format="$ %,.0f"),
				"Presupuesto_Escala_2": st.column_config.NumberColumn("Escala 2", format="$ %,.0f"),
				"Cumplimiento_E1": st.column_config.ProgressColumn("Cumpl. E1", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Cumplimiento_E2": st.column_config.ProgressColumn("Cumpl. E2", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Rebate_Mensual_Ganado": st.column_config.NumberColumn("Rebate mensual", format="$ %,.0f"),
				"Bono_Estacionalidad": st.column_config.NumberColumn("Bono estacionalidad", format="$ %,.0f"),
				"Pendiente_Cartera": st.column_config.NumberColumn("Pendiente cartera", format="$ %,.0f"),
			},
		)

	with monthly_tab:
		st.subheader("Seguimiento mensual")
		monthly_display_df = monthly_df[["Mes", "Trimestre", "Compra_Aplicable", "Presupuesto_Escala_1", "Presupuesto_Escala_2", "Cumplimiento_E1", "Cumplimiento_E2", "Faltante_E1", "Faltante_E2", "Escala_Lograda", "Estado_Mes", "Compra_Hasta_Corte", "Meta_Estacionalidad", "Avance_Estacionalidad", "Estado_Estacionalidad", "Rebate_Mensual_Ganado", "Bono_Estacionalidad", "Pendiente_Cartera"]]
		st.dataframe(
			monthly_display_df,
			width="stretch",
			hide_index=True,
			column_config={
				"Compra_Aplicable": st.column_config.NumberColumn("Compra aplicable", format="$ %,.0f"),
				"Presupuesto_Escala_1": st.column_config.NumberColumn("Meta E1", format="$ %,.0f"),
				"Presupuesto_Escala_2": st.column_config.NumberColumn("Meta E2", format="$ %,.0f"),
				"Cumplimiento_E1": st.column_config.ProgressColumn("Cumpl. E1", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Cumplimiento_E2": st.column_config.ProgressColumn("Cumpl. E2", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Faltante_E1": st.column_config.NumberColumn("Faltante E1", format="$ %,.0f"),
				"Faltante_E2": st.column_config.NumberColumn("Faltante E2", format="$ %,.0f"),
				"Compra_Hasta_Corte": st.column_config.NumberColumn("Compra al corte", format="$ %,.0f"),
				"Meta_Estacionalidad": st.column_config.NumberColumn("Meta estacionalidad", format="$ %,.0f"),
				"Avance_Estacionalidad": st.column_config.ProgressColumn("Avance estacionalidad", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Rebate_Mensual_Ganado": st.column_config.NumberColumn("Rebate mensual", format="$ %,.0f"),
				"Bono_Estacionalidad": st.column_config.NumberColumn("Bono estacionalidad", format="$ %,.0f"),
				"Pendiente_Cartera": st.column_config.NumberColumn("Pendiente cartera", format="$ %,.0f"),
			},
		)
		st.bar_chart(monthly_df.set_index("Mes")[["Rebate_Mensual_Ganado", "Bono_Estacionalidad"]], width="stretch")

	with quarter_tab:
		st.subheader("Cumplimiento trimestral y recuperación")
		st.dataframe(
			quarterly_df,
			width="stretch",
			hide_index=True,
			column_config={
				"Compra_Aplicable": st.column_config.NumberColumn("Compra aplicable", format="$ %,.0f"),
				"Presupuesto_Escala_1": st.column_config.NumberColumn("Meta E1", format="$ %,.0f"),
				"Presupuesto_Escala_2": st.column_config.NumberColumn("Meta E2", format="$ %,.0f"),
				"Cumplimiento_E1": st.column_config.ProgressColumn("Cumpl. E1", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Cumplimiento_E2": st.column_config.ProgressColumn("Cumpl. E2", min_value=0.0, max_value=1.2, format="%.0f%%"),
				"Faltante_E1": st.column_config.NumberColumn("Faltante E1", format="$ %,.0f"),
				"Faltante_E2": st.column_config.NumberColumn("Faltante E2", format="$ %,.0f"),
				"Rebate_Trimestral_Ganado": st.column_config.NumberColumn("Rebate trimestral", format="$ %,.0f"),
				"Recomposicion_Trimestral_Proyectada": st.column_config.NumberColumn("Recuperable", format="$ %,.0f"),
				"Recomposicion_Cartera_Bloqueada": st.column_config.NumberColumn("Bloqueado cartera", format="$ %,.0f"),
			},
		)
		st.bar_chart(quarterly_df.set_index("Trimestre")[["Rebate_Trimestral_Ganado", "Recomposicion_Trimestral_Proyectada", "Recomposicion_Cartera_Bloqueada"]], width="stretch")
		st.markdown(
			f"""
			<div class="section-card">
				<strong>Recomposición 9 meses</strong><br><br>
				Escala acumulada al ritmo actual: {pill_html(cycle_outlook['Escala_Ciclo'], get_scale_tone(cycle_outlook['Escala_Ciclo']))}<br><br>
				Recuperable proyectado: <strong>{format_currency(cycle_outlook['Recomposicion_9M_Proyectada'])}</strong><br>
				Bloqueado por cartera: <strong>{format_currency(cycle_outlook['Recomposicion_9M_Bloqueada'])}</strong><br>
				Faltante ciclo a E2: <strong>{format_currency(cycle_outlook['Faltante_E2_Ciclo'])}</strong>
			</div>
			""",
			unsafe_allow_html=True,
		)

	with invoices_tab:
		render_provider_invoice_tab(provider_config, pintuco_df, snapshot_date, show_exclusion_columns=True)

	with diagnostics_tab:
		executive_export_df = pd.DataFrame([
			{
				"Foto": datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M:%S"),
				"Compra_Aplicable_Acumulada": cycle_outlook["Compra_Aplicable_Acumulada"],
				"Meta_Acumulada_E1": cycle_outlook["Meta_Acumulada_E1"],
				"Meta_Acumulada_E2": cycle_outlook["Meta_Acumulada_E2"],
				"Escala_Ciclo": cycle_outlook["Escala_Ciclo"],
				"Cumplimiento_Acumulado_E1": cycle_outlook["Cumplimiento_Acumulado_E1"],
				"Cumplimiento_Acumulado_E2": cycle_outlook["Cumplimiento_Acumulado_E2"],
				"Rebate_Ganado_A_Hoy": rebate_total_ganado,
				"Recomposicion_9M_Proyectada": cycle_outlook["Recomposicion_9M_Proyectada"],
				"Recomposicion_9M_Bloqueada": cycle_outlook["Recomposicion_9M_Bloqueada"],
			}
		])
		render_provider_diagnostics_tab(
			provider_config["label"],
			pintuco_df,
			tracking_alerts,
			[("Resumen_Ejecutivo", executive_export_df), ("Mes_a_Mes", monthly_df), ("Trimestres", quarterly_df), ("Facturas_Pintuco", pintuco_df)],
			f"Rebate_Pintuco_Ejecutivo_{snapshot_date}.xlsx",
		)


provider_configs = get_provider_configs()
provider_tabs = st.tabs(["🎨 Pintuco", "🧱 Abracol", "🟦 Goya"])

with provider_tabs[0]:
	render_pintuco_dashboard(provider_configs["pintuco"])

with provider_tabs[1]:
	render_abracol_dashboard(provider_configs["abracol"])

with provider_tabs[2]:
	render_goya_dashboard(provider_configs["goya"])
