# -*- coding: utf-8 -*-
"""Core operativo para conciliacion, tesoreria y programacion de pagos."""

import email
import imaplib
import io
import os
import re
import uuid
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime, timedelta
from email.header import decode_header
from email.utils import parsedate_to_datetime
from typing import Any, Optional

import dropbox
import gspread
import numpy as np
import pandas as pd
import pytz
import requests
import streamlit as st
from google.oauth2.service_account import Credentials
from openpyxl.utils import get_column_letter
from streamlit.errors import StreamlitSecretNotFoundError


COLOMBIA_TZ = pytz.timezone("America/Bogota")
IMAP_SERVER = "imap.gmail.com"
EMAIL_FOLDER = "TFHKA/Recepcion/Descargados"
DROPBOX_PENDING_PATH = "/data/Proveedores.csv"
DROPBOX_PAID_PATH = "/data/cartera_saldada.csv"
PROVIDER_CATALOG_PATH = "PROVEDORES_CORREO.xlsx"

SHEET_PROVIDER_MASTER = "Maestro_Proveedores"
SHEET_EMAIL_HISTORY = "Historial_Correo_Proveedores"
SHEET_MASTER_INVOICES = "Maestro_Facturas"
SHEET_PAYMENT_PLAN = "Propuesta_Pagos"
SHEET_PAYMENT_LOTS = "Lotes_Pago"
SHEET_EMAIL_LOG = "Historial_Correos"

PENDING_COLUMNS = [
    "nombre_proveedor_erp",
    "serie",
    "num_entrada",
    "num_factura",
    "doc_erp",
    "fecha_emision_erp",
    "fecha_vencimiento_erp",
    "valor_total_erp",
]

PAID_COLUMNS = [
    "nombre_proveedor_erp",
    "serie",
    "num_entrada",
    "num_factura",
    "estado_documento",
    "fecha_emision_erp",
    "fecha_vencimiento_erp",
    "valor_total_erp",
]

EMAIL_COLUMNS = [
    "invoice_key",
    "num_factura",
    "proveedor_correo",
    "proveedor_norm",
    "fecha_emision_correo",
    "fecha_vencimiento_correo",
    "valor_total_correo",
    "fecha_recepcion_correo",
    "remitente_correo",
    "asunto_correo",
    "nombre_adjunto",
    "message_id",
]

PROVIDER_MASTER_COLUMNS = [
    "codigo_proveedor",
    "nif",
    "proveedor",
    "proveedor_norm",
    "activo",
    "email_pago",
    "email_cc",
    "email_alertas",
    "contacto_pagos",
    "contacto_tesoreria",
    "telefono",
    "condiciones_comerciales",
    "observaciones",
]

PAYMENT_LOT_COLUMNS = [
    "lote_id",
    "fecha_registro",
    "fecha_programada_pago",
    "responsable",
    "invoice_key",
    "proveedor",
    "num_factura",
    "valor_factura",
    "valor_descuento",
    "valor_a_pagar",
    "estado_lote",
    "motivo_pago",
    "email_destino",
]

EMAIL_LOG_COLUMNS = [
    "envio_id",
    "fecha_envio",
    "lote_id",
    "proveedor",
    "email_destino",
    "email_cc",
    "asunto",
    "facturas",
    "ahorro_total",
    "estado_envio",
    "detalle_envio",
]

MASTER_OPTIONAL_DEFAULTS = {
    "invoice_key": "",
    "proveedor": "",
    "proveedor_norm": "",
    "proveedor_erp": "",
    "proveedor_correo": "",
    "num_factura": "",
    "estado_erp": "No ERP",
    "estado_conciliacion": "Sin clasificar",
    "estado_vencimiento": "No aplica",
    "valor_erp": 0.0,
    "valor_total_correo": 0.0,
    "diferencia_valor": 0.0,
    "detalle_valor": "",
    "detalle_conciliacion": "",
    "valor_descuento": 0.0,
    "valor_a_pagar": 0.0,
    "descuento_pct": 0.0,
    "fecha_limite_descuento": pd.NaT,
    "fecha_vencimiento_erp": pd.NaT,
    "fecha_emision_erp": pd.NaT,
    "fecha_emision_correo": pd.NaT,
    "fecha_vencimiento_correo": pd.NaT,
    "fecha_recepcion_correo": pd.NaT,
    "fecha_programada_pago": pd.NaT,
    "estado_descuento": "No aplica",
    "registrada_para_pago": False,
    "riesgo_mora_48h": False,
    "dias_para_vencer": 0,
    "remitente_correo": "",
    "asunto_correo": "",
    "nombre_adjunto": "",
    "message_id": "",
    "motivo_base": "",
    "lote_id": "",
    "estado_lote": "",
    "email_pago": "",
    "email_cc": "",
    "email_alertas": "",
    "contacto_pagos": "",
    "condiciones_comerciales": "",
    "activo": True,
}

_NUMERIC_COLS = ["valor_erp", "valor_total_correo", "diferencia_valor", "valor_descuento", "valor_a_pagar", "descuento_pct", "dias_para_vencer"]
_DATETIME_COLS = ["fecha_limite_descuento", "fecha_vencimiento_erp", "fecha_emision_erp", "fecha_emision_correo", "fecha_vencimiento_correo", "fecha_recepcion_correo", "fecha_programada_pago"]
_BOOLEAN_COLS = ["registrada_para_pago", "riesgo_mora_48h"]

DISCOUNT_PROVIDERS = {
    "ABRACOL S.A.S": [{"days": 8, "rate": 0.04}],
    "ASSA ABLOY COLOMBIA S.A.S": [{"days": 8, "rate": 0.03}],
    "DELTA GLOBAL S.A.S": [{"days": 10, "rate": 0.03}],
    "INDUMA S.C.A": [{"days": 10, "rate": 0.025}, {"days": 30, "rate": 0.01}],
    "INDUSTRIAS GOYAINCOL LTDA": [{"days": 30, "rate": 0.05}],
    "INDUSTRIAS GOYAINCOL SAS": [{"days": 30, "rate": 0.05}],
    "ARTECOLA COLOMBIA S.A.S": [{"days": 10, "rate": 0.025}],
    "PINTUCO COLOMBIA S.A.S": [{"days": 15, "rate": 0.03}, {"days": 30, "rate": 0.02}],
    "SAINT - GOBAIN COLOMBIA S.A.S.": [{"days": 10, "rate": 0.025}, {"days": 20, "rate": 0.015}, {"days": 30, "rate": 0.007}],
    "RODILLOS MASTDER S.A.S": [{"days": 10, "rate": 0.05}],
    "SEGUREX LATAM S.A.S": [{"days": 8, "rate": 0.03}],
}
DISCOUNT_RULES_NORMALIZED = {
    re.sub(r"[^A-Z0-9]", "", provider.upper()): rules for provider, rules in DISCOUNT_PROVIDERS.items()
}
SUPPLIER_ALIASES = {
    "COMPANIAGLOBALDEPINTURASSAS": "PINTUCOCOLOMBIASAS",
    "COMPANIAGLOBALDEPINTURASSASS": "PINTUCOCOLOMBIASAS",
    "PINTUCOSAC": "PINTUCOCOLOMBIASAS",
    "INDUSTRIASGOYAINCOLLTDA": "INDUSTRIASGOYAINCOLSAS",
}
VALUE_MATCH_RULES = {
    "INDUSTRIASGOYAINCOLSAS": {"retention_pct": 0.025},
}


def get_secrets_dict() -> dict[str, Any]:
    try:
        return st.secrets.to_dict()
    except StreamlitSecretNotFoundError:
        return {}
    except Exception:
        return {}


def get_secret_value(key: str, default: Any = None) -> Any:
    return get_secrets_dict().get(key, default)


def get_secret_section(key: str) -> dict[str, Any]:
    section = get_secret_value(key, {})
    if isinstance(section, dict):
        return section
    try:
        return dict(section)
    except Exception:
        return {}


def ensure_authenticated() -> None:
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    if not st.session_state["password_correct"]:
        st.error("Debes iniciar sesion desde Dashboard General para acceder a esta pagina.")
        st.stop()


def format_currency(value: Any) -> str:
    return f"${float(value or 0):,.0f}"


def normalize_text(value: Any) -> str:
    if pd.isna(value) or value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).upper()).strip()


def normalize_supplier_key(value: Any) -> str:
    normalized = normalize_text(value)
    return SUPPLIER_ALIASES.get(normalized, normalized)


def normalize_supplier_fingerprint(value: Any) -> str:
    if pd.isna(value) or value is None:
        return ""
    tokens = [token for token in re.findall(r"[A-Z0-9]+", str(value).upper()) if token]
    return "".join(sorted(tokens))


def normalize_invoice_number(value: Any) -> str:
    return normalize_text(value)


def clean_numeric(value: Any) -> float:
    if pd.isna(value) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).strip().replace("$", "").replace("COP", "").replace(" ", "")
    if "." in cleaned and "," in cleaned:
        if cleaned.rfind(".") > cleaned.rfind(","):
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(",", "")

    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def coerce_datetime(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    if not pd.api.types.is_datetime64_any_dtype(parsed):
        return pd.Series(pd.NaT, index=series.index)
    try:
        if getattr(parsed.dt, "tz", None) is None:
            return parsed.dt.tz_localize(COLOMBIA_TZ, ambiguous="infer")
        return parsed.dt.tz_convert(COLOMBIA_TZ)
    except (TypeError, ValueError, AttributeError):
        return pd.Series(pd.NaT, index=series.index)


def coalesce(*series_list: pd.Series) -> pd.Series:
    if not series_list:
        return pd.Series(dtype=object)
    result = series_list[0].copy()
    for series in series_list[1:]:
        result = result.where(result.notna() & (result.astype(str) != ""), series)
    return result


def prepare_invoice_key(df: pd.DataFrame, provider_col: str, invoice_col: str = "num_factura") -> pd.DataFrame:
    prepared = df.copy()
    prepared[provider_col] = prepared[provider_col].fillna("").astype(str)
    prepared["proveedor_norm"] = prepared[provider_col].apply(normalize_supplier_key)
    prepared[invoice_col] = prepared[invoice_col].apply(normalize_invoice_number)
    prepared["invoice_key"] = prepared["proveedor_norm"] + "|" + prepared[invoice_col]
    return prepared


def first_non_empty_value(series: pd.Series) -> Any:
    valid = series.dropna().astype(str).str.strip()
    valid = valid[valid.ne("") & valid.ne("nan") & valid.ne("None")]
    if not valid.empty:
        return valid.iloc[0]
    non_null = series.dropna()
    return non_null.iloc[0] if not non_null.empty else ""


def aggregate_erp_invoice_rows(df: pd.DataFrame, source_label: str) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    prepared = df.copy()
    if "invoice_key" not in prepared.columns:
        provider_col = "nombre_proveedor_erp" if "nombre_proveedor_erp" in prepared.columns else "proveedor_correo"
        prepared = prepare_invoice_key(prepared, provider_col)

    agg_map: dict[str, Any] = {
        "nombre_proveedor_erp": first_non_empty_value,
        "serie": first_non_empty_value,
        "num_entrada": first_non_empty_value,
        "num_factura": first_non_empty_value,
        "fecha_emision_erp": "min",
        "fecha_vencimiento_erp": "max",
        "valor_total_erp": "sum",
        "proveedor_norm": first_non_empty_value,
    }
    if "doc_erp" in prepared.columns:
        agg_map["doc_erp"] = first_non_empty_value
    if "estado_documento" in prepared.columns:
        agg_map["estado_documento"] = first_non_empty_value

    agg_map = {key: value for key, value in agg_map.items() if key in prepared.columns}
    aggregated = prepared.groupby("invoice_key", dropna=False, as_index=False).agg(agg_map)
    counts = prepared.groupby("invoice_key", dropna=False).size().reset_index(name="erp_movimientos")
    aggregated = aggregated.merge(counts, on="invoice_key", how="left")
    aggregated["estado_erp_fuente"] = source_label
    return aggregated


def build_provider_matching_maps(provider_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    if provider_df.empty:
        return {
            "nif_to_norm": {},
            "fingerprint_to_norm": {},
        }

    prepared = provider_df.copy()
    if "proveedor_norm" not in prepared.columns and "proveedor" in prepared.columns:
        prepared["proveedor_norm"] = prepared["proveedor"].apply(normalize_supplier_key)

    prepared["supplier_fingerprint"] = prepared.get("proveedor", pd.Series(index=prepared.index, dtype=object)).apply(normalize_supplier_fingerprint)
    prepared["nif_norm"] = prepared.get("nif", pd.Series(index=prepared.index, dtype=object)).apply(normalize_text)

    nif_to_norm: dict[str, str] = {}
    for _, row in prepared.iterrows():
        nif_norm = str(row.get("nif_norm", "") or "")
        proveedor_norm = str(row.get("proveedor_norm", "") or "")
        if nif_norm and proveedor_norm and nif_norm not in nif_to_norm:
            nif_to_norm[nif_norm] = proveedor_norm

    fingerprint_counts = prepared.groupby("supplier_fingerprint")["proveedor_norm"].nunique().to_dict()
    fingerprint_to_norm: dict[str, str] = {}
    for _, row in prepared.iterrows():
        fingerprint = str(row.get("supplier_fingerprint", "") or "")
        proveedor_norm = str(row.get("proveedor_norm", "") or "")
        if fingerprint and proveedor_norm and fingerprint_counts.get(fingerprint) == 1:
            fingerprint_to_norm[fingerprint] = proveedor_norm

    return {
        "nif_to_norm": nif_to_norm,
        "fingerprint_to_norm": fingerprint_to_norm,
    }


def align_email_records_to_erp(
    email_df: pd.DataFrame,
    pending_df: pd.DataFrame,
    paid_df: pd.DataFrame,
) -> pd.DataFrame:
    if email_df.empty:
        return email_df

    erp_frames = []
    for frame in [pending_df, paid_df]:
        if frame.empty:
            continue
        provider_col = "nombre_proveedor_erp" if "nombre_proveedor_erp" in frame.columns else "proveedor_correo"
        prepared = frame.copy()
        if "invoice_key" not in prepared.columns:
            prepared = prepare_invoice_key(prepared, provider_col)
        prepared["num_factura"] = prepared["num_factura"].apply(normalize_invoice_number)
        prepared["supplier_fingerprint"] = prepared.get(provider_col, pd.Series(index=prepared.index, dtype=object)).apply(normalize_supplier_fingerprint)
        erp_frames.append(prepared[["invoice_key", "num_factura", "proveedor_norm", "supplier_fingerprint"]])

    if not erp_frames:
        return email_df

    erp_df = pd.concat(erp_frames, ignore_index=True).drop_duplicates(subset=["invoice_key"], keep="first")
    erp_keys = set(erp_df["invoice_key"].astype(str).tolist())

    aligned = email_df.copy()
    aligned["email_supplier_fingerprint"] = aligned.get("proveedor_correo", pd.Series(index=aligned.index, dtype=object)).apply(normalize_supplier_fingerprint)

    for idx, row in aligned.iterrows():
        current_key = str(row.get("invoice_key", "") or "")
        invoice_number = normalize_invoice_number(row.get("num_factura", ""))
        if not invoice_number or current_key in erp_keys:
            continue

        candidates = erp_df[erp_df["num_factura"] == invoice_number].copy()
        if candidates.empty:
            continue

        provider_norm = str(row.get("proveedor_norm", "") or "")
        if provider_norm:
            exact_candidates = candidates[candidates["proveedor_norm"] == provider_norm]
            if not exact_candidates.empty:
                candidates = exact_candidates

        if len(candidates) > 1:
            fingerprint = str(row.get("email_supplier_fingerprint", "") or "")
            if fingerprint:
                fingerprint_candidates = candidates[candidates["supplier_fingerprint"] == fingerprint]
                if len(fingerprint_candidates) == 1:
                    candidates = fingerprint_candidates

        if len(candidates) == 1:
            aligned.at[idx, "proveedor_norm"] = candidates["proveedor_norm"].iloc[0]
            aligned.at[idx, "invoice_key"] = candidates["invoice_key"].iloc[0]

    return aligned.drop(columns=["email_supplier_fingerprint"], errors="ignore")


@st.cache_resource(show_spinner="Conectando a Google Sheets...")
def connect_to_google_sheets() -> Optional[gspread.Client]:
    try:
        google_credentials = get_secret_value("google_credentials")
        if not google_credentials:
            return None
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(google_credentials, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as exc:
        st.error(f"❌ Error al autenticar con Google Sheets: {exc}")
        return None


def get_or_create_worksheet(client: gspread.Client, worksheet_name: str) -> gspread.Worksheet:
    sheet_id = get_secret_value("google_sheet_id")
    if not sheet_id:
        raise ValueError("No existe google_sheet_id en los secretos configurados.")
    spreadsheet = client.open_by_key(sheet_id)
    try:
        return spreadsheet.worksheet(worksheet_name)
    except gspread.WorksheetNotFound:
        return spreadsheet.add_worksheet(title=worksheet_name, rows="4000", cols="80")


def load_sheet_df(client: gspread.Client, worksheet_name: str) -> pd.DataFrame:
    try:
        worksheet = get_or_create_worksheet(client, worksheet_name)
        records = worksheet.get_all_records()
        return pd.DataFrame(records) if records else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def update_worksheet_from_df(worksheet: gspread.Worksheet, df: pd.DataFrame) -> bool:
    try:
        df_to_upload = df.copy()
        for column in df_to_upload.columns:
            if pd.api.types.is_datetime64_any_dtype(df_to_upload[column]):
                df_to_upload[column] = pd.to_datetime(df_to_upload[column], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")

        # Fill all NaN/NaT/None BEFORE converting to str to avoid float('nan') leaking into JSON
        df_to_upload = df_to_upload.fillna("").astype(str).replace({"nan": "", "NaT": "", "None": "", "<NA>": "", "inf": "", "-inf": ""})
        data = [df_to_upload.columns.tolist()] + df_to_upload.values.tolist() if not df_to_upload.empty else [df_to_upload.columns.tolist()]
        worksheet.update(data, "A1")

        total_rows = worksheet.row_count
        used_rows = len(data)
        if used_rows < total_rows and df_to_upload.columns.tolist():
            last_col = get_column_letter(len(df_to_upload.columns))
            empty_rows = [[""] * len(df_to_upload.columns) for _ in range(total_rows - used_rows)]
            worksheet.update(f"A{used_rows + 1}:{last_col}{total_rows}", empty_rows, raw=False)
        return True
    except Exception as exc:
        st.error(f"❌ Error actualizando la hoja {worksheet.title}: {exc}")
        return False


def save_df_to_sheet(client: gspread.Client, worksheet_name: str, df: pd.DataFrame) -> bool:
    worksheet = get_or_create_worksheet(client, worksheet_name)
    return update_worksheet_from_df(worksheet, df)


def append_df_to_sheet(client: gspread.Client, worksheet_name: str, new_rows_df: pd.DataFrame, ordered_columns: list[str]) -> bool:
    existing = load_sheet_df(client, worksheet_name)
    combined = pd.concat([existing, new_rows_df], ignore_index=True) if not existing.empty else new_rows_df.copy()
    for column in ordered_columns:
        if column not in combined.columns:
            combined[column] = ""
    combined = combined.reindex(columns=ordered_columns)
    return save_df_to_sheet(client, worksheet_name, combined)


def decode_mime_text(value: Any) -> str:
    if not value:
        return ""
    decoded = []
    for part, encoding in decode_header(value):
        if isinstance(part, bytes):
            decoded.append(part.decode(encoding or "utf-8", errors="ignore"))
        else:
            decoded.append(part)
    return "".join(decoded).strip()


def parse_email_datetime(value: str) -> pd.Timestamp:
    if not value:
        return pd.NaT
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = COLOMBIA_TZ.localize(dt)
        else:
            dt = dt.astimezone(COLOMBIA_TZ)
        return pd.Timestamp(dt)
    except Exception:
        return pd.NaT


def load_provider_master_base() -> pd.DataFrame:
    if not os.path.exists(PROVIDER_CATALOG_PATH):
        st.error(f"No se encontro el archivo {PROVIDER_CATALOG_PATH} en la raiz del proyecto.")
        return pd.DataFrame(columns=PROVIDER_MASTER_COLUMNS)

    base_df = pd.read_excel(PROVIDER_CATALOG_PATH)
    provider_col = next((col for col in base_df.columns if "proveedor" in str(col).lower()), None)
    code_col = next((col for col in base_df.columns if "código" in str(col).lower() or "codigo" in str(col).lower()), None)
    nif_col = next((col for col in base_df.columns if "nif" in str(col).lower() or "nit" in str(col).lower()), None)
    if provider_col is None:
        st.error("PROVEDORES_CORREO.xlsx debe contener una columna de proveedor.")
        return pd.DataFrame(columns=PROVIDER_MASTER_COLUMNS)

    provider_df = pd.DataFrame(
        {
            "codigo_proveedor": base_df[code_col] if code_col else "",
            "nif": base_df[nif_col] if nif_col else "",
            "proveedor": base_df[provider_col].fillna("").astype(str).str.strip(),
        }
    )
    provider_df["proveedor_norm"] = provider_df["proveedor"].apply(normalize_supplier_key)
    provider_df["activo"] = True
    provider_df["email_pago"] = ""
    provider_df["email_cc"] = ""
    provider_df["email_alertas"] = ""
    provider_df["contacto_pagos"] = ""
    provider_df["contacto_tesoreria"] = ""
    provider_df["telefono"] = ""
    provider_df["condiciones_comerciales"] = ""
    provider_df["observaciones"] = ""
    return provider_df[PROVIDER_MASTER_COLUMNS].drop_duplicates(subset=["proveedor_norm"], keep="first")


def load_provider_master(client: gspread.Client) -> pd.DataFrame:
    base_df = load_provider_master_base()
    sheet_df = load_sheet_df(client, SHEET_PROVIDER_MASTER)

    if sheet_df.empty:
        return base_df

    for column in PROVIDER_MASTER_COLUMNS:
        if column not in sheet_df.columns:
            sheet_df[column] = "" if column != "activo" else True

    base_df = base_df.set_index("proveedor_norm")
    sheet_df = sheet_df.set_index("proveedor_norm")
    editable_cols = [
        "activo",
        "email_pago",
        "email_cc",
        "email_alertas",
        "contacto_pagos",
        "contacto_tesoreria",
        "telefono",
        "condiciones_comerciales",
        "observaciones",
    ]

    merged = base_df.combine_first(sheet_df)
    for column in editable_cols:
        if column in sheet_df.columns:
            merged[column] = sheet_df[column].where(sheet_df[column].astype(str) != "", merged[column])

    merged.reset_index(inplace=True)
    merged["activo"] = merged["activo"].astype(str).str.lower().map({"false": False, "0": False}).fillna(True)
    return merged[PROVIDER_MASTER_COLUMNS].drop_duplicates(subset=["proveedor_norm"], keep="first")


def save_provider_master(client: gspread.Client, provider_df: pd.DataFrame) -> bool:
    prepared = provider_df.copy()
    for column in PROVIDER_MASTER_COLUMNS:
        if column not in prepared.columns:
            prepared[column] = "" if column != "activo" else True
    return save_df_to_sheet(client, SHEET_PROVIDER_MASTER, prepared[PROVIDER_MASTER_COLUMNS])


@st.cache_data(ttl=600, show_spinner="Descargando cartera pendiente desde Dropbox...")
def load_pending_invoices_from_dropbox() -> pd.DataFrame:
    try:
        dropbox_secrets = get_secret_section("dropbox")
        if not dropbox_secrets:
            return pd.DataFrame(columns=PENDING_COLUMNS + ["proveedor_norm", "invoice_key", "estado_erp_fuente"])
        dbx = dropbox.Dropbox(
            oauth2_refresh_token=dropbox_secrets.get("refresh_token"),
            app_key=dropbox_secrets.get("app_key"),
            app_secret=dropbox_secrets.get("app_secret"),
        )
        _, response = dbx.files_download(DROPBOX_PENDING_PATH)
        df = pd.read_csv(
            io.StringIO(response.content.decode("latin1")),
            sep="{",
            header=None,
            engine="python",
            names=PENDING_COLUMNS,
        )
        df["valor_total_erp"] = df["valor_total_erp"].apply(clean_numeric)
        df["fecha_emision_erp"] = coerce_datetime(df["fecha_emision_erp"])
        df["fecha_vencimiento_erp"] = coerce_datetime(df["fecha_vencimiento_erp"])
        df.dropna(subset=["nombre_proveedor_erp"], inplace=True)

        credit_mask = (df["valor_total_erp"] < 0) & (
            df["num_factura"].isna() | (df["num_factura"].astype(str).str.strip() == "")
        )
        if credit_mask.any():
            df.loc[credit_mask, "num_factura"] = (
                "NC-"
                + df.loc[credit_mask, "doc_erp"].astype(str).str.strip()
                + "-"
                + df.loc[credit_mask, "valor_total_erp"].abs().astype(int).astype(str)
            )

        df = prepare_invoice_key(df, "nombre_proveedor_erp")
        return aggregate_erp_invoice_rows(df, "Pendiente")
    except Exception as exc:
        st.error(f"❌ Error cargando cartera pendiente desde Dropbox: {exc}")
        return pd.DataFrame(columns=PENDING_COLUMNS + ["proveedor_norm", "invoice_key", "estado_erp_fuente"])


@st.cache_data(ttl=600, show_spinner="Descargando cartera saldada desde Dropbox...")
def load_paid_invoices_from_dropbox() -> pd.DataFrame:
    try:
        dropbox_secrets = get_secret_section("dropbox")
        if not dropbox_secrets:
            return pd.DataFrame(columns=PAID_COLUMNS + ["proveedor_norm", "invoice_key", "estado_erp_fuente"])

        paid_path = dropbox_secrets.get("paid_invoices_path") or dropbox_secrets.get("paid_path") or DROPBOX_PAID_PATH
        dbx = dropbox.Dropbox(
            oauth2_refresh_token=dropbox_secrets.get("refresh_token"),
            app_key=dropbox_secrets.get("app_key"),
            app_secret=dropbox_secrets.get("app_secret"),
        )
        _, response = dbx.files_download(paid_path)
        df = pd.read_csv(
            io.StringIO(response.content.decode("latin1")),
            sep="|",
            encoding="latin1",
            header=None,
            names=PAID_COLUMNS,
            engine="python",
        )
        df["valor_total_erp"] = df["valor_total_erp"].apply(clean_numeric)
        df["fecha_emision_erp"] = coerce_datetime(df["fecha_emision_erp"])
        df["fecha_vencimiento_erp"] = coerce_datetime(df["fecha_vencimiento_erp"])
        df = prepare_invoice_key(df, "nombre_proveedor_erp")
        df["doc_erp"] = ""
        return aggregate_erp_invoice_rows(df, "Saldada")
    except Exception as exc:
        st.error(f"❌ Error cargando cartera saldada desde Dropbox: {exc}")
        return pd.DataFrame(columns=PAID_COLUMNS + ["proveedor_norm", "invoice_key", "estado_erp_fuente"])


def parse_invoice_xml(xml_content: str, target_suppliers: set[str], provider_maps: Optional[dict[str, dict[str, str]]] = None) -> Optional[dict]:
    try:
        namespaces = {
            "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
            "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
        }
        xml_content = re.sub(r"^[^<]+", "", xml_content.strip())
        root = ET.fromstring(xml_content.encode("utf-8"))

        description_node = root.find(".//cac:Attachment/cac:ExternalReference/cbc:Description", namespaces)
        if description_node is not None and description_node.text and "<Invoice" in description_node.text:
            nested_xml = re.sub(r"^[^<]+", "", description_node.text.strip())
            root = ET.fromstring(nested_xml.encode("utf-8"))

        def find_text(paths: list[str]) -> Optional[str]:
            for path in paths:
                node = root.find(path, namespaces)
                if node is not None and node.text:
                    return node.text.strip()
            return None

        supplier_name = find_text([
            ".//cac:AccountingSupplierParty/cac:Party/cac:PartyName/cbc:Name",
            ".//cac:AccountingSupplierParty/cac:Party/cac:PartyLegalEntity/cbc:RegistrationName",
        ])
        supplier_nif = find_text([
            ".//cac:AccountingSupplierParty/cac:Party/cac:PartyTaxScheme/cbc:CompanyID",
            ".//cac:AccountingSupplierParty/cac:Party/cac:PartyLegalEntity/cbc:CompanyID",
        ])
        invoice_number = find_text(["./cbc:ID"])
        issue_date = find_text(["./cbc:IssueDate"])
        due_date = find_text(["./cbc:DueDate", ".//cac:PaymentMeans/cbc:PaymentDueDate"])
        total_value = find_text([".//cac:LegalMonetaryTotal/cbc:PayableAmount", ".//cac:LegalMonetaryTotal/cbc:TaxExclusiveAmount"])

        if not supplier_name or not invoice_number or not total_value:
            return None

        supplier_norm = normalize_supplier_key(supplier_name)
        supplier_nif_norm = normalize_text(supplier_nif)
        supplier_fingerprint = normalize_supplier_fingerprint(supplier_name)

        if provider_maps:
            supplier_norm = provider_maps.get("nif_to_norm", {}).get(supplier_nif_norm, supplier_norm)
            if supplier_norm not in target_suppliers:
                supplier_norm = provider_maps.get("fingerprint_to_norm", {}).get(supplier_fingerprint, supplier_norm)

        if target_suppliers and supplier_norm not in target_suppliers:
            return None

        return {
            "proveedor_correo": supplier_name,
            "proveedor_norm": supplier_norm,
            "supplier_nif": supplier_nif_norm,
            "num_factura": normalize_invoice_number(invoice_number),
            "fecha_emision_correo": issue_date,
            "fecha_vencimiento_correo": due_date,
            "valor_total_correo": clean_numeric(total_value),
        }
    except Exception:
        return None


def extract_invoice_records_from_message(message_obj, target_suppliers: set[str], provider_maps: Optional[dict[str, dict[str, str]]] = None) -> tuple[list[dict], dict]:
    records = []
    stats = {"attachments_scanned": 0, "xml_files_scanned": 0, "invoice_rows_detected": 0}

    email_subject = decode_mime_text(message_obj.get("Subject", ""))
    email_sender = decode_mime_text(message_obj.get("From", ""))
    email_message_id = decode_mime_text(message_obj.get("Message-ID", ""))
    email_received_at = parse_email_datetime(message_obj.get("Date", ""))

    for part in message_obj.walk():
        if part.get_content_maintype() == "multipart":
            continue

        filename = decode_mime_text(part.get_filename() or "")
        content_type = (part.get_content_type() or "").lower()
        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        is_zip = filename.lower().endswith(".zip") or content_type in {"application/zip", "application/x-zip-compressed"}
        is_xml = filename.lower().endswith(".xml") or content_type in {"application/xml", "text/xml"}
        if not is_zip and not is_xml:
            continue

        stats["attachments_scanned"] += 1
        if is_zip:
            try:
                with zipfile.ZipFile(io.BytesIO(payload)) as zip_file:
                    for internal_name in zip_file.namelist():
                        if not internal_name.lower().endswith(".xml"):
                            continue
                        stats["xml_files_scanned"] += 1
                        xml_content = zip_file.read(internal_name).decode("utf-8", "ignore")
                        details = parse_invoice_xml(xml_content, target_suppliers, provider_maps)
                        if details:
                            details.update(
                                {
                                    "fecha_recepcion_correo": email_received_at,
                                    "remitente_correo": email_sender,
                                    "asunto_correo": email_subject,
                                    "nombre_adjunto": internal_name,
                                    "message_id": email_message_id,
                                }
                            )
                            records.append(details)
                            stats["invoice_rows_detected"] += 1
            except Exception:
                continue

        if is_xml:
            stats["xml_files_scanned"] += 1
            xml_content = payload.decode("utf-8", "ignore")
            details = parse_invoice_xml(xml_content, target_suppliers, provider_maps)
            if details:
                details.update(
                    {
                        "fecha_recepcion_correo": email_received_at,
                        "remitente_correo": email_sender,
                        "asunto_correo": email_subject,
                        "nombre_adjunto": filename or "adjunto_xml",
                        "message_id": email_message_id,
                    }
                )
                records.append(details)
                stats["invoice_rows_detected"] += 1

    return records, stats


def fetch_supplier_invoices_from_email(start_date: date, target_suppliers: set[str], provider_df: Optional[pd.DataFrame] = None) -> tuple[pd.DataFrame, dict]:
    invoices_data = []
    stats = {
        "emails_found": 0,
        "emails_processed": 0,
        "attachments_scanned": 0,
        "xml_files_scanned": 0,
        "invoice_rows_detected": 0,
        "started_from": start_date.strftime("%Y-%m-%d"),
    }
    provider_maps = build_provider_matching_maps(provider_df if provider_df is not None else pd.DataFrame())

    try:
        email_secrets = get_secret_section("email")
        if not email_secrets.get("address") or not email_secrets.get("password"):
            return pd.DataFrame(columns=EMAIL_COLUMNS), stats
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(email_secrets["address"], email_secrets["password"])
        mail.select(f'"{EMAIL_FOLDER}"')
        _, messages = mail.search(None, f'(SINCE "{start_date.strftime("%d-%b-%Y")}")')
        message_ids = messages[0].split()
        stats["emails_found"] = len(message_ids)
        if not message_ids:
            mail.logout()
            return pd.DataFrame(columns=EMAIL_COLUMNS), stats

        progress_bar = st.progress(0, text=f"Procesando {len(message_ids)} correos de proveedores...")
        for index, message_id in enumerate(message_ids, start=1):
            _, data = mail.fetch(message_id, "(RFC822)")
            message_obj = email.message_from_bytes(data[0][1])
            records, email_stats = extract_invoice_records_from_message(message_obj, target_suppliers, provider_maps)
            invoices_data.extend(records)

            stats["emails_processed"] += 1
            stats["attachments_scanned"] += email_stats["attachments_scanned"]
            stats["xml_files_scanned"] += email_stats["xml_files_scanned"]
            stats["invoice_rows_detected"] += email_stats["invoice_rows_detected"]
            progress_bar.progress(index / len(message_ids), text=f"Procesando {index}/{len(message_ids)} correos...")

        mail.logout()
    except Exception as exc:
        st.error(f"❌ Error procesando correos de proveedores: {exc}")

    email_df = pd.DataFrame(invoices_data)
    if email_df.empty:
        return pd.DataFrame(columns=EMAIL_COLUMNS), stats

    email_df = prepare_invoice_key(email_df, "proveedor_correo")
    email_df["fecha_emision_correo"] = coerce_datetime(email_df["fecha_emision_correo"])
    email_df["fecha_vencimiento_correo"] = coerce_datetime(email_df["fecha_vencimiento_correo"])
    email_df["fecha_recepcion_correo"] = coerce_datetime(email_df["fecha_recepcion_correo"])
    email_df["valor_total_correo"] = email_df["valor_total_correo"].apply(clean_numeric)

    for column in EMAIL_COLUMNS:
        if column not in email_df.columns:
            email_df[column] = ""
    email_df = email_df[EMAIL_COLUMNS].drop_duplicates(subset=["invoice_key", "message_id"], keep="last")
    return email_df, stats


def determine_sync_start(email_history_df: pd.DataFrame) -> date:
    if email_history_df.empty or "fecha_recepcion_correo" not in email_history_df.columns:
        return date(datetime.now(COLOMBIA_TZ).year, 1, 1)
    history = email_history_df.copy()
    history["fecha_recepcion_correo"] = coerce_datetime(history["fecha_recepcion_correo"])
    last_date = history["fecha_recepcion_correo"].max()
    if pd.isna(last_date):
        return date(datetime.now(COLOMBIA_TZ).year, 1, 1)
    return max(date(datetime.now(COLOMBIA_TZ).year, 1, 1), (last_date - timedelta(days=3)).date())


def merge_email_history(existing_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    if existing_df.empty:
        merged = new_df.copy()
    else:
        merged = pd.concat([existing_df, new_df], ignore_index=True)
    for column in EMAIL_COLUMNS:
        if column not in merged.columns:
            merged[column] = ""
    merged["fecha_recepcion_correo"] = coerce_datetime(merged["fecha_recepcion_correo"])
    merged = merged[EMAIL_COLUMNS].drop_duplicates(subset=["invoice_key", "message_id"], keep="last")
    return merged.sort_values(by=["fecha_recepcion_correo", "invoice_key"], ascending=False)


def apply_discount_rules(master_df: pd.DataFrame) -> pd.DataFrame:
    df = master_df.copy()
    df["descuento_pct"] = 0.0
    df["valor_descuento"] = 0.0
    df["valor_a_pagar"] = df["valor_erp"].fillna(0.0)
    df["fecha_limite_descuento"] = pd.Series([pd.NaT] * len(df), index=df.index, dtype="object")
    df["estado_descuento"] = "No aplica"

    today = pd.Timestamp.now(tz=COLOMBIA_TZ).normalize()
    for index, row in df.iterrows():
        if row.get("estado_erp") != "Pendiente" or row.get("valor_erp", 0) <= 0 or pd.isna(row.get("fecha_emision_erp")):
            continue
        provider_rules = DISCOUNT_RULES_NORMALIZED.get(row.get("proveedor_norm", ""), [])
        if not provider_rules:
            continue

        valid_rules = []
        for rule in provider_rules:
            deadline = row["fecha_emision_erp"] + timedelta(days=rule["days"])
            if deadline.normalize() >= today:
                valid_rules.append((rule["rate"], deadline))

        if not valid_rules:
            df.at[index, "estado_descuento"] = "Ventana vencida"
            continue

        valid_rules.sort(key=lambda item: (-item[0], item[1]))
        rate, deadline = valid_rules[0]
        discount_value = row["valor_erp"] * rate

        df.at[index, "descuento_pct"] = rate
        df.at[index, "valor_descuento"] = discount_value
        df.at[index, "valor_a_pagar"] = row["valor_erp"] - discount_value
        df.at[index, "fecha_limite_descuento"] = deadline
        df.at[index, "estado_descuento"] = f"Disponible {rate:.1%}"

    df["fecha_limite_descuento"] = coerce_datetime(df["fecha_limite_descuento"])
    return df


def evaluate_value_status(row: pd.Series) -> tuple[bool, str]:
    if not row.get("en_correo") or row.get("estado_erp") not in {"Pendiente", "Saldada"}:
        return False, ""

    value_erp = clean_numeric(row.get("valor_erp"))
    value_email = clean_numeric(row.get("valor_total_correo"))
    difference = abs(value_erp - value_email)
    if difference <= 1:
        return False, ""

    provider_rule = VALUE_MATCH_RULES.get(str(row.get("proveedor_norm", "")), {})
    retention_pct = provider_rule.get("retention_pct")
    if retention_pct and value_email > 0:
        expected_erp_value = value_email * (1 - retention_pct)
        tolerance = max(100.0, value_email * 0.003)
        if abs(value_erp - expected_erp_value) <= tolerance:
            return False, f"Diferencia compatible con retencion {retention_pct:.1%}"

    return True, "Diferencia de valor por revisar"


def build_master_dataframe(
    pending_df: pd.DataFrame,
    paid_df: pd.DataFrame,
    email_df: pd.DataFrame,
    provider_df: pd.DataFrame,
    lot_history_df: pd.DataFrame,
) -> pd.DataFrame:
    pending_df = pending_df.copy()
    paid_df = paid_df.copy()
    email_df = email_df.copy()

    # Ensure invoice_key exists on every frame (fix: reassign back, handle empty)
    if not pending_df.empty and "invoice_key" not in pending_df.columns:
        provider_col = "nombre_proveedor_erp" if "nombre_proveedor_erp" in pending_df.columns else "proveedor_correo"
        pending_df = prepare_invoice_key(pending_df, provider_col)
    if not paid_df.empty and "invoice_key" not in paid_df.columns:
        provider_col = "nombre_proveedor_erp" if "nombre_proveedor_erp" in paid_df.columns else "proveedor_correo"
        paid_df = prepare_invoice_key(paid_df, provider_col)
    if not email_df.empty and "invoice_key" not in email_df.columns:
        provider_col = "proveedor_correo" if "proveedor_correo" in email_df.columns else "nombre_proveedor_erp"
        email_df = prepare_invoice_key(email_df, provider_col)

    # Guarantee invoice_key column even on empty frames
    for frame in [pending_df, paid_df, email_df]:
        if "invoice_key" not in frame.columns:
            frame["invoice_key"] = pd.Series(dtype=str)

    email_df = align_email_records_to_erp(email_df, pending_df, paid_df)

    pending_df = pending_df.drop_duplicates(subset=["invoice_key"], keep="last")
    paid_df = paid_df.drop_duplicates(subset=["invoice_key"], keep="last")
    email_df = email_df.drop_duplicates(subset=["invoice_key"], keep="last")

    combined = pd.merge(pending_df, paid_df, on="invoice_key", how="outer", suffixes=("_pend", "_paid"))
    combined = pd.merge(combined, email_df, on="invoice_key", how="outer")

    _provider_cols = [c for c in ["proveedor_norm", "proveedor", "email_pago", "email_cc", "email_alertas", "contacto_pagos", "condiciones_comerciales", "activo"] if c in provider_df.columns]
    provider_meta = provider_df[_provider_cols].drop_duplicates(subset=["proveedor_norm"]) if "proveedor_norm" in provider_df.columns and not provider_df.empty else pd.DataFrame(columns=["proveedor_norm"])
    combined["proveedor_norm"] = coalesce(
        combined.get("proveedor_norm_pend", pd.Series(index=combined.index)),
        combined.get("proveedor_norm_paid", pd.Series(index=combined.index)),
        combined.get("proveedor_norm", pd.Series(index=combined.index)),
    )
    combined["proveedor_norm"] = combined["proveedor_norm"].fillna("").astype(str).replace({"nan": "", "None": ""})
    provider_meta["proveedor_norm"] = provider_meta["proveedor_norm"].fillna("").astype(str)
    combined = combined.merge(provider_meta, on="proveedor_norm", how="left")

    lot_latest = pd.DataFrame(columns=["invoice_key", "lote_id", "estado_lote", "fecha_programada_pago"])
    if not lot_history_df.empty:
        lot_latest = lot_history_df.copy()
        if "fecha_registro" in lot_latest.columns:
            lot_latest["fecha_registro"] = coerce_datetime(lot_latest["fecha_registro"])
            lot_latest.sort_values(by="fecha_registro", inplace=True)
        lot_latest = lot_latest.drop_duplicates(subset=["invoice_key"], keep="last")

    combined = combined.merge(lot_latest[[col for col in ["invoice_key", "lote_id", "estado_lote", "fecha_programada_pago"] if col in lot_latest.columns]], on="invoice_key", how="left")

    combined["num_factura"] = coalesce(
        combined.get("num_factura_pend", pd.Series(index=combined.index)),
        combined.get("num_factura_paid", pd.Series(index=combined.index)),
        combined.get("num_factura", pd.Series(index=combined.index)),
    )
    combined["proveedor_erp"] = coalesce(
        combined.get("nombre_proveedor_erp_pend", pd.Series(index=combined.index)),
        combined.get("nombre_proveedor_erp_paid", pd.Series(index=combined.index)),
    )
    combined["proveedor_correo"] = combined.get("proveedor_correo", pd.Series(index=combined.index)).fillna("")
    combined["proveedor"] = coalesce(combined.get("proveedor", pd.Series(index=combined.index)), combined["proveedor_erp"], combined["proveedor_correo"])

    combined["fecha_emision_erp"] = coalesce(
        combined.get("fecha_emision_erp_pend", pd.Series(index=combined.index)),
        combined.get("fecha_emision_erp_paid", pd.Series(index=combined.index)),
    )
    combined["fecha_vencimiento_erp"] = coalesce(
        combined.get("fecha_vencimiento_erp_pend", pd.Series(index=combined.index)),
        combined.get("fecha_vencimiento_erp_paid", pd.Series(index=combined.index)),
    )
    combined["valor_erp"] = coalesce(
        combined.get("valor_total_erp_pend", pd.Series(index=combined.index)),
        combined.get("valor_total_erp_paid", pd.Series(index=combined.index)),
    ).apply(clean_numeric)
    combined["valor_total_correo"] = combined.get("valor_total_correo", pd.Series(index=combined.index)).apply(clean_numeric)

    combined["fecha_emision_erp"] = coerce_datetime(combined["fecha_emision_erp"])
    combined["fecha_vencimiento_erp"] = coerce_datetime(combined["fecha_vencimiento_erp"])
    combined["fecha_emision_correo"] = coerce_datetime(combined.get("fecha_emision_correo", pd.Series(index=combined.index)))
    combined["fecha_vencimiento_correo"] = coerce_datetime(combined.get("fecha_vencimiento_correo", pd.Series(index=combined.index)))
    combined["fecha_recepcion_correo"] = coerce_datetime(combined.get("fecha_recepcion_correo", pd.Series(index=combined.index)))
    combined["fecha_programada_pago"] = coerce_datetime(combined.get("fecha_programada_pago", pd.Series(index=combined.index)))

    combined["en_pendiente"] = combined["num_factura_pend"].notna() if "num_factura_pend" in combined.columns else False
    combined["en_saldada"] = combined["num_factura_paid"].notna() if "num_factura_paid" in combined.columns else False
    combined["movimiento_mixto_erp"] = combined["en_pendiente"] & combined["en_saldada"]
    combined["en_correo"] = combined["num_factura"].notna() & combined["proveedor_correo"].astype(str).ne("")

    # Determine which invoices are older than the email reading window.
    # The email sync starts from Jan 1 of the current year, so invoices
    # emitted before that date will never have email support found.
    email_window_start = pd.Timestamp(date(datetime.now(COLOMBIA_TZ).year, 1, 1), tz=COLOMBIA_TZ)
    _fe = combined["fecha_emision_erp"].copy()
    # Use emission date; fall back to due date if emission is missing
    _fe = _fe.fillna(combined["fecha_vencimiento_erp"])
    combined["anterior_a_lectura_correo"] = _fe.notna() & (_fe < email_window_start)

    conditions = [
        combined["en_pendiente"],
        combined["en_saldada"],
    ]
    choices = ["Pendiente", "Saldada"]
    combined["estado_erp"] = pd.Series(pd.NA, index=combined.index, dtype="object")
    combined.loc[conditions[0], "estado_erp"] = choices[0]
    combined.loc[conditions[1] & combined["estado_erp"].isna(), "estado_erp"] = choices[1]
    combined["estado_erp"] = combined["estado_erp"].fillna("No ERP")

    combined["diferencia_valor"] = (combined["valor_erp"].fillna(0) - combined["valor_total_correo"].fillna(0)).abs()
    value_status = combined.apply(evaluate_value_status, axis=1)
    combined["tiene_discrepancia_valor"] = value_status.apply(lambda item: item[0])
    combined["detalle_valor"] = value_status.apply(lambda item: item[1])

    def classify_status(row: pd.Series) -> str:
        if row["tiene_discrepancia_valor"] and row["estado_erp"] == "Pendiente":
            return "Pendiente con valor por revisar"
        if row["tiene_discrepancia_valor"] and row["estado_erp"] == "Saldada":
            return "Saldada con valor por revisar"
        if row["estado_erp"] == "Pendiente" and row["en_correo"]:
            return "Pendiente conciliada"
        if row["estado_erp"] == "Pendiente" and not row["en_correo"]:
            # If the invoice was emitted before the email reading window,
            # it is expected not to have email support — not a real gap.
            if row.get("anterior_a_lectura_correo", False):
                return "Pendiente anterior a lectura"
            return "Pendiente sin correo"
        if row["estado_erp"] == "Saldada" and row["en_correo"]:
            return "Saldada conciliada"
        if row["estado_erp"] == "Saldada" and not row["en_correo"]:
            # Old paid invoices without email are expected — not a gap.
            if row.get("anterior_a_lectura_correo", False):
                return "Saldada anterior a lectura"
            return "Saldada sin correo"
        if row["estado_erp"] == "No ERP" and row["en_correo"]:
            return "Solo correo"
        return "Sin clasificar"

    def build_reconciliation_detail(row: pd.Series) -> str:
        if row["estado_conciliacion"] == "Pendiente conciliada":
            if row.get("movimiento_mixto_erp", False):
                return "Existe en cartera pendiente, tiene soporte de correo y además presenta movimientos saldados en ERP. Se toma como pendiente por saldo abierto."
            if row.get("detalle_valor"):
                return f"Existe en cartera pendiente y tiene soporte de correo. {row['detalle_valor']}."
            return "Existe en cartera pendiente y tiene soporte de correo conciliado."
        if row["estado_conciliacion"] == "Saldada conciliada":
            if row.get("detalle_valor"):
                return f"Existe en cartera saldada y tiene soporte de correo. {row['detalle_valor']}."
            return "Existe en cartera saldada y tiene soporte de correo conciliado."
        if row["estado_conciliacion"] == "Pendiente sin correo":
            if row.get("movimiento_mixto_erp", False):
                return "Está en cartera pendiente y además tiene movimientos saldados en ERP, pero no se encontró soporte de correo para el saldo abierto."
            return "Está en cartera pendiente pero no se encontró XML o ZIP asociado en el buzón leído."
        if row["estado_conciliacion"] == "Pendiente anterior a lectura":
            return "Está en cartera pendiente. La fecha de emisión es anterior a la ventana de lectura de correo; no aplica cruce documental."
        if row["estado_conciliacion"] == "Saldada sin correo":
            return "Está en cartera saldada pero no se encontró XML o ZIP asociado en el buzón leído."
        if row["estado_conciliacion"] == "Saldada anterior a lectura":
            return "Está en cartera saldada. La fecha de emisión es anterior a la ventana de lectura de correo; no aplica cruce documental."
        if row["estado_conciliacion"] == "Solo correo":
            return "Tiene correo y XML, pero no aparece en las fuentes ERP descargadas desde Dropbox (cartera pendiente ni cartera saldada)."
        if row["estado_conciliacion"] in {"Pendiente con valor por revisar", "Saldada con valor por revisar"}:
            return row.get("detalle_valor") or "Existe soporte de correo, pero el valor no coincide con ERP."
        return "Revisar manualmente este caso."

    combined["estado_conciliacion"] = combined.apply(classify_status, axis=1)
    combined["detalle_conciliacion"] = combined.apply(build_reconciliation_detail, axis=1)
    today = pd.Timestamp.now(tz=COLOMBIA_TZ).normalize()
    combined["dias_para_vencer"] = (combined["fecha_vencimiento_erp"].dt.normalize() - today).dt.days

    def classify_due(row: pd.Series) -> str:
        if row["estado_erp"] != "Pendiente":
            return "No aplica"
        if pd.isna(row["fecha_vencimiento_erp"]):
            return "Sin fecha ERP"
        if row["dias_para_vencer"] < 0:
            return "🔴 Vencida"
        if row["dias_para_vencer"] <= 2:
            return "🟠 Riesgo 48h"
        if row["dias_para_vencer"] <= 7:
            return "🟡 Proxima a vencer"
        return "🟢 Vigente"

    combined["estado_vencimiento"] = combined.apply(classify_due, axis=1)
    combined["riesgo_mora_48h"] = combined["estado_vencimiento"].eq("🟠 Riesgo 48h")
    combined["registrada_para_pago"] = combined.get("lote_id", pd.Series(index=combined.index)).fillna("").astype(str).ne("")
    combined["motivo_base"] = combined.apply(
        lambda row: "Pendiente con soporte de correo" if row["estado_conciliacion"] == "Pendiente conciliada"
        else "Pendiente sin soporte recibido" if row["estado_conciliacion"] == "Pendiente sin correo"
        else "Correo sin reflejo en fuentes ERP Dropbox" if row["estado_conciliacion"] == "Solo correo"
        else "Saldo abierto con movimientos mixtos ERP" if row.get("movimiento_mixto_erp", False)
        else "Factura ya saldada" if row["estado_erp"] == "Saldada"
        else "Revisar caso manualmente",
        axis=1,
    )

    combined = apply_discount_rules(combined)
    combined["prioridad_descuento"] = combined["descuento_pct"].fillna(0)

    selected_columns = [
        "invoice_key",
        "num_factura",
        "proveedor",
        "proveedor_norm",
        "proveedor_erp",
        "proveedor_correo",
        "valor_erp",
        "valor_total_correo",
        "diferencia_valor",
        "detalle_valor",
        "fecha_emision_erp",
        "fecha_vencimiento_erp",
        "fecha_emision_correo",
        "fecha_vencimiento_correo",
        "fecha_recepcion_correo",
        "remitente_correo",
        "asunto_correo",
        "nombre_adjunto",
        "message_id",
        "estado_erp",
        "estado_conciliacion",
        "detalle_conciliacion",
        "estado_vencimiento",
        "dias_para_vencer",
        "riesgo_mora_48h",
        "descuento_pct",
        "valor_descuento",
        "valor_a_pagar",
        "fecha_limite_descuento",
        "estado_descuento",
        "motivo_base",
        "lote_id",
        "estado_lote",
        "fecha_programada_pago",
        "registrada_para_pago",
        "email_pago",
        "email_cc",
        "email_alertas",
        "contacto_pagos",
        "condiciones_comerciales",
        "activo",
    ]
    for column in selected_columns:
        if column not in combined.columns:
            combined[column] = ""

    combined = combined[selected_columns].drop_duplicates(subset=["invoice_key"], keep="first")
    combined.sort_values(by=["proveedor", "fecha_vencimiento_erp", "num_factura"], inplace=True)
    return combined


def build_payment_plan(master_df: pd.DataFrame) -> pd.DataFrame:
    if master_df.empty:
        return pd.DataFrame()

    plan_df = master_df[(master_df["estado_erp"] == "Pendiente") & (master_df["valor_erp"] > 0)].copy()
    if plan_df.empty:
        return pd.DataFrame()

    today = pd.Timestamp.now(tz=COLOMBIA_TZ).normalize()
    plan_df["fecha_objetivo"] = plan_df["fecha_limite_descuento"].fillna(plan_df["fecha_vencimiento_erp"])
    plan_df["dias_para_objetivo"] = (plan_df["fecha_objetivo"].dt.normalize() - today).dt.days

    def assign_reason(row: pd.Series) -> str:
        if row["descuento_pct"] > 0 and pd.notna(row["fecha_limite_descuento"]):
            if row["dias_para_objetivo"] <= 2:
                return "Asegurar pronto pago antes de perder descuento"
            return "Capturar descuento financiero disponible"
        if row["riesgo_mora_48h"]:
            return "Evitar mora en las proximas 48 horas"
        if row["estado_vencimiento"] == "🔴 Vencida":
            return "Regularizar factura vencida"
        return "Programar dentro de ventana normal"

    plan_df["motivo_pago"] = plan_df.apply(assign_reason, axis=1)
    plan_df["ranking_descuento"] = -plan_df["descuento_pct"].fillna(0)
    plan_df["ranking_vencimiento"] = plan_df["dias_para_objetivo"].fillna(9999)
    plan_df["ranking_valor"] = -plan_df["valor_descuento"].fillna(0)
    plan_df.sort_values(by=["ranking_vencimiento", "ranking_descuento", "ranking_valor", "proveedor"], inplace=True)
    plan_df["prioridad_pago"] = range(1, len(plan_df) + 1)

    display_columns = [
        "prioridad_pago",
        "invoice_key",
        "proveedor",
        "num_factura",
        "valor_erp",
        "descuento_pct",
        "valor_descuento",
        "valor_a_pagar",
        "fecha_vencimiento_erp",
        "fecha_limite_descuento",
        "dias_para_vencer",
        "dias_para_objetivo",
        "estado_vencimiento",
        "estado_conciliacion",
        "motivo_pago",
        "lote_id",
        "estado_lote",
        "email_pago",
        "email_cc",
    ]
    return plan_df[[c for c in display_columns if c in plan_df.columns]]


def build_risk_alerts(master_df: pd.DataFrame) -> pd.DataFrame:
    if master_df.empty:
        return pd.DataFrame()
    alerts_df = master_df[(master_df.get("estado_erp", pd.Series(dtype=object)) == "Pendiente") & (master_df.get("estado_vencimiento", pd.Series(dtype=object)).isin(["🟠 Riesgo 48h", "🔴 Vencida"]))].copy()
    if alerts_df.empty:
        return pd.DataFrame()
    alerts_df["tipo_alerta"] = alerts_df["estado_vencimiento"].map({"🟠 Riesgo 48h": "Riesgo de mora 48h", "🔴 Vencida": "Factura vencida"})
    _acols = [c for c in ["invoice_key", "proveedor", "num_factura", "valor_erp", "fecha_vencimiento_erp", "dias_para_vencer", "tipo_alerta", "email_alertas"] if c in alerts_df.columns]
    _asort = [c for c in ["dias_para_vencer", "proveedor"] if c in alerts_df.columns]
    return alerts_df[_acols].sort_values(by=_asort) if _asort else alerts_df[_acols]


def ensure_master_dataframe_schema(master_df: pd.DataFrame) -> pd.DataFrame:
    if master_df.empty:
        return master_df.copy()

    prepared = master_df.copy()
    for column, default in MASTER_OPTIONAL_DEFAULTS.items():
        if column not in prepared.columns:
            prepared[column] = default

    for col in _NUMERIC_COLS:
        if col in prepared.columns:
            prepared[col] = prepared[col].apply(clean_numeric)

    for col in _DATETIME_COLS:
        if col in prepared.columns:
            prepared[col] = coerce_datetime(prepared[col])

    for col in _BOOLEAN_COLS:
        if col in prepared.columns:
            prepared[col] = (
                prepared[col].astype(str).str.strip().str.lower()
                .map({"true": True, "1": True, "1.0": True, "false": False, "0": False, "0.0": False})
                .fillna(False)
            )

    if (prepared["valor_a_pagar"] == 0).all() and "valor_erp" in prepared.columns:
        prepared["valor_a_pagar"] = prepared["valor_erp"] - prepared["valor_descuento"]

    return prepared


def safe_display(
    df: pd.DataFrame,
    columns: list[str],
    sort_by: Optional[list[str]] = None,
    ascending: Any = True,
) -> pd.DataFrame:
    """Safely select columns and sort a DataFrame for display without KeyError."""
    if df.empty:
        return pd.DataFrame(columns=[c for c in columns if c in df.columns])
    existing_sort = [c for c in (sort_by or []) if c in df.columns]
    if existing_sort:
        if isinstance(ascending, list) and sort_by:
            asc_map = dict(zip(sort_by, ascending))
            safe_asc = [asc_map.get(c, True) for c in existing_sort]
        else:
            safe_asc = ascending
        result = df.sort_values(by=existing_sort, ascending=safe_asc)
    else:
        result = df
    existing_cols = [c for c in columns if c in result.columns]
    return result[existing_cols] if existing_cols else pd.DataFrame()


def infer_payload_snapshot_metadata(payload: dict) -> dict[str, Any]:
    timestamp_candidates = []
    candidate_columns = [
        (payload.get("master_df", pd.DataFrame()), ["fecha_recepcion_correo", "fecha_programada_pago", "fecha_vencimiento_erp"]),
        (payload.get("email_history_df", pd.DataFrame()), ["fecha_recepcion_correo"]),
        (payload.get("lot_history_df", pd.DataFrame()), ["fecha_registro", "fecha_programada_pago"]),
        (payload.get("email_log_df", pd.DataFrame()), ["fecha_envio"]),
    ]

    for frame, columns in candidate_columns:
        if frame is None or frame.empty:
            continue
        for column in columns:
            if column in frame.columns:
                parsed = coerce_datetime(frame[column])
                latest = parsed.max()
                if pd.notna(latest):
                    timestamp_candidates.append(latest)

    snapshot_at = max(timestamp_candidates) if timestamp_candidates else pd.NaT
    has_snapshot = any(
        not payload.get(key, pd.DataFrame()).empty
        for key in ["master_df", "payment_plan_df", "risk_alerts_df", "email_history_df"]
    )
    snapshot_rows = len(payload.get("master_df", pd.DataFrame())) if not payload.get("master_df", pd.DataFrame()).empty else len(payload.get("email_history_df", pd.DataFrame()))

    return {
        "has_snapshot": has_snapshot,
        "snapshot_rows": snapshot_rows,
        "snapshot_at": snapshot_at,
    }


def sync_treasury_data() -> dict:
    """Sincronización progresiva: guarda cada paso de inmediato para no perder avance."""
    gs_client = connect_to_google_sheets()
    if not gs_client:
        st.error("No fue posible conectar con Google Sheets.")
        return {}

    sync_errors: list[str] = []
    sync_stats: dict = {}
    step = st.status("Sincronizando fuentes...", expanded=True)

    # ── 1. Proveedores y datos ya guardados en Sheets ──────────────
    step.update(label="Cargando maestro de proveedores y datos previos...")
    provider_df = load_provider_master(gs_client)
    target_suppliers = set(
        provider_df[provider_df["activo"].fillna(True)]["proveedor_norm"].tolist()
    )
    email_history_df = load_sheet_df(gs_client, SHEET_EMAIL_HISTORY)
    lot_history_df = load_sheet_df(gs_client, SHEET_PAYMENT_LOTS)
    email_log_df = load_sheet_df(gs_client, SHEET_EMAIL_LOG)
    save_provider_master(gs_client, provider_df)
    st.write(f"✅ Proveedores: {len(provider_df):,} | Historial correo previo: {len(email_history_df):,}")

    # ── 2. Dropbox: cartera pendiente y saldada (rápido, ~segundos) ─
    step.update(label="Descargando cartera desde Dropbox...")
    pending_df = pd.DataFrame()
    paid_df = pd.DataFrame()
    try:
        pending_df = load_pending_invoices_from_dropbox()
        paid_df = load_paid_invoices_from_dropbox()
        if target_suppliers:
            pending_df = pending_df[pending_df["proveedor_norm"].isin(target_suppliers)].copy()
            paid_df = paid_df[paid_df["proveedor_norm"].isin(target_suppliers)].copy()
        st.write(f"✅ Pendientes Dropbox: {len(pending_df):,} | Saldadas Dropbox: {len(paid_df):,}")
    except Exception as exc:
        sync_errors.append(f"Dropbox: {exc}")
        st.warning(f"⚠️ Error cargando Dropbox: {exc}")

    # ── 3. Guardar foto parcial solo con Dropbox (si hay datos) ─────
    #    Así al reabrir la app ya se ve cartera aunque el correo falle.
    merged_email_df = email_history_df  # lo que ya existía
    try:
        partial_master = build_master_dataframe(
            pending_df, paid_df, merged_email_df, provider_df, lot_history_df
        )
        save_df_to_sheet(gs_client, SHEET_MASTER_INVOICES, partial_master)
        st.write(f"✅ Foto parcial guardada: {len(partial_master):,} registros (Dropbox + historial previo)")
    except Exception as exc:
        sync_errors.append(f"Foto parcial: {exc}")
        st.warning(f"⚠️ No se pudo guardar foto parcial: {exc}")
        partial_master = pd.DataFrame()

    # ── 4. Correo: solo la ventana incremental ──────────────────────
    step.update(label="Leyendo correos nuevos (incremental)...")
    start_date = determine_sync_start(email_history_df)
    st.write(f"📧 Leyendo correos desde **{start_date}** (incremental)")
    try:
        new_email_df, sync_stats = fetch_supplier_invoices_from_email(
            start_date, target_suppliers, provider_df
        )
        merged_email_df = merge_email_history(email_history_df, new_email_df)
        # Guardar historial de correo de inmediato para no perder progreso
        save_df_to_sheet(gs_client, SHEET_EMAIL_HISTORY, merged_email_df)
        st.write(
            f"✅ Correos procesados: {sync_stats.get('emails_processed', 0):,} | "
            f"XML encontrados: {sync_stats.get('invoice_rows_detected', 0):,} | "
            f"Historial total: {len(merged_email_df):,}"
        )
    except Exception as exc:
        sync_errors.append(f"Correo: {exc}")
        st.warning(f"⚠️ Error leyendo correo: {exc}. Se conserva historial previo ({len(merged_email_df):,} registros).")

    # ── 5. Construir maestro final con toda la información ──────────
    step.update(label="Construyendo maestro consolidado...")
    try:
        master_df = build_master_dataframe(
            pending_df, paid_df, merged_email_df, provider_df, lot_history_df
        )
        payment_plan_df = build_payment_plan(master_df)
        risk_alerts_df = build_risk_alerts(master_df)
        save_df_to_sheet(gs_client, SHEET_MASTER_INVOICES, master_df)
        save_df_to_sheet(gs_client, SHEET_PAYMENT_PLAN, payment_plan_df)
        st.write(
            f"✅ Maestro final: {len(master_df):,} | "
            f"Plan de pagos: {len(payment_plan_df):,} | "
            f"Alertas: {len(risk_alerts_df):,}"
        )
    except Exception as exc:
        sync_errors.append(f"Maestro: {exc}")
        st.warning(f"⚠️ Error construyendo maestro final: {exc}")
        master_df = partial_master if not partial_master.empty else pd.DataFrame()
        master_df = ensure_master_dataframe_schema(master_df)
        payment_plan_df = build_payment_plan(master_df)
        risk_alerts_df = build_risk_alerts(master_df)
        # Guardar aunque sea parcial
        try:
            save_df_to_sheet(gs_client, SHEET_MASTER_INVOICES, master_df)
        except Exception:
            pass

    # ── 6. Resultado ────────────────────────────────────────────────
    if sync_errors:
        step.update(label="Sincronización completada con advertencias", state="error")
        for err in sync_errors:
            st.error(f"⛔ {err}")
    else:
        step.update(label="Sincronización completada correctamente", state="complete")

    payload = {
        "provider_df": provider_df,
        "email_history_df": merged_email_df,
        "pending_df": pending_df,
        "paid_df": paid_df,
        "master_df": master_df,
        "payment_plan_df": payment_plan_df,
        "risk_alerts_df": risk_alerts_df,
        "lot_history_df": lot_history_df,
        "email_log_df": email_log_df,
        "sync_stats": sync_stats,
        "sync_started_from": start_date.strftime("%Y-%m-%d"),
        "snapshot_source": "live_sync",
    }

    payload.update(infer_payload_snapshot_metadata(payload))

    st.session_state["treasury_payload"] = payload
    st.session_state["last_treasury_sync"] = datetime.now(COLOMBIA_TZ).strftime("%Y-%m-%d %H:%M:%S")
    return payload


def load_operational_payload() -> dict:
    """Carga la foto guardada en Google Sheets. No toca correo ni Dropbox."""
    payload = st.session_state.get("treasury_payload")
    if payload:
        return payload

    gs_client = connect_to_google_sheets()
    if not gs_client:
        return {}

    try:
        provider_df = load_provider_master(gs_client)
    except Exception:
        provider_df = pd.DataFrame()

    try:
        master_df = ensure_master_dataframe_schema(
            load_sheet_df(gs_client, SHEET_MASTER_INVOICES)
        )
    except Exception:
        master_df = pd.DataFrame()

    try:
        email_history_df = load_sheet_df(gs_client, SHEET_EMAIL_HISTORY)
    except Exception:
        email_history_df = pd.DataFrame()

    try:
        lot_history_df = load_sheet_df(gs_client, SHEET_PAYMENT_LOTS)
    except Exception:
        lot_history_df = pd.DataFrame()

    try:
        email_log_df = load_sheet_df(gs_client, SHEET_EMAIL_LOG)
    except Exception:
        email_log_df = pd.DataFrame()

    payload = {
        "provider_df": provider_df,
        "email_history_df": email_history_df,
        "master_df": master_df,
        "payment_plan_df": build_payment_plan(master_df),
        "risk_alerts_df": build_risk_alerts(master_df),
        "lot_history_df": lot_history_df,
        "email_log_df": email_log_df,
        "snapshot_source": "sheets_cache",
    }
    payload.update(infer_payload_snapshot_metadata(payload))
    st.session_state["treasury_payload"] = payload
    return payload


def build_payment_email_html(provider_name: str, lot_df: pd.DataFrame, payment_date: date, notes: str = "") -> str:
    savings = lot_df["valor_descuento"].sum()
    total_to_pay = lot_df["valor_a_pagar"].sum()
    rows_html = "".join(
        [
            f"""
            <tr>
                <td style='padding:10px 12px;border-bottom:1px solid #e5edf5;'>{row['num_factura']}</td>
                <td style='padding:10px 12px;border-bottom:1px solid #e5edf5;text-align:right;'>{format_currency(row['valor_erp'])}</td>
                <td style='padding:10px 12px;border-bottom:1px solid #e5edf5;text-align:right;'>{format_currency(row['valor_descuento'])}</td>
                <td style='padding:10px 12px;border-bottom:1px solid #e5edf5;text-align:right;'>{format_currency(row['valor_a_pagar'])}</td>
            </tr>
            """
            for _, row in lot_df.iterrows()
        ]
    )

    notes_block = f"<p style='margin-top:18px;color:#506070;font-size:14px;'>{notes}</p>" if notes else ""
    return f"""
    <div style="font-family:Helvetica,Arial,sans-serif;background:#f4f7fb;padding:24px;">
        <div style="max-width:860px;margin:0 auto;background:#ffffff;border-radius:24px;overflow:hidden;border:1px solid #dbe5f0;box-shadow:0 18px 44px rgba(12,45,87,.08);">
            <div style="background:linear-gradient(120deg,#0c2d57 0%,#195b97 60%,#f0a202 100%);padding:28px 32px;color:#ffffff;">
                <div style="font-size:12px;letter-spacing:.08em;text-transform:uppercase;opacity:.9;">Ferreinox S.A.S. BIC</div>
                <h1 style="margin:8px 0 0 0;font-size:30px;line-height:1.1;">Programacion de pago propuesta</h1>
                <p style="margin:10px 0 0 0;font-size:15px;opacity:.94;">Proveedor: <strong>{provider_name}</strong></p>
            </div>
            <div style="padding:28px 32px;">
                <p style="font-size:15px;color:#334155;line-height:1.6;">Compartimos el lote de facturas programadas para pago con fecha objetivo <strong>{payment_date.strftime('%Y-%m-%d')}</strong>. Esta propuesta prioriza la conservacion de descuentos financieros y la prevencion de mora.</p>
                <div style="display:flex;gap:12px;flex-wrap:wrap;margin:18px 0 22px 0;">
                    <div style="background:#f8fbff;border:1px solid #dce8f5;border-radius:16px;padding:14px 16px;min-width:180px;">
                        <div style="font-size:12px;color:#6b7c8f;text-transform:uppercase;">Facturas</div>
                        <div style="font-size:26px;font-weight:800;color:#0c2d57;">{len(lot_df)}</div>
                    </div>
                    <div style="background:#f8fbff;border:1px solid #dce8f5;border-radius:16px;padding:14px 16px;min-width:180px;">
                        <div style="font-size:12px;color:#6b7c8f;text-transform:uppercase;">Descuento ganado</div>
                        <div style="font-size:26px;font-weight:800;color:#116149;">{format_currency(savings)}</div>
                    </div>
                    <div style="background:#f8fbff;border:1px solid #dce8f5;border-radius:16px;padding:14px 16px;min-width:180px;">
                        <div style="font-size:12px;color:#6b7c8f;text-transform:uppercase;">Valor a pagar</div>
                        <div style="font-size:26px;font-weight:800;color:#0c2d57;">{format_currency(total_to_pay)}</div>
                    </div>
                </div>
                <table style="width:100%;border-collapse:collapse;border:1px solid #e5edf5;border-radius:14px;overflow:hidden;">
                    <thead>
                        <tr style="background:#f5f9fd;color:#0c2d57;text-align:left;">
                            <th style='padding:12px;'>Factura</th>
                            <th style='padding:12px;text-align:right;'>Valor original</th>
                            <th style='padding:12px;text-align:right;'>Descuento</th>
                            <th style='padding:12px;text-align:right;'>Valor a pagar</th>
                        </tr>
                    </thead>
                    <tbody>{rows_html}</tbody>
                </table>
                {notes_block}
                <p style="margin-top:22px;font-size:14px;color:#516173;line-height:1.6;">Agradecemos confirmar cualquier novedad documental o financiera sobre este lote. Este correo fue generado desde el centro de tesoreria de Ferreinox para mantener trazabilidad operativa.</p>
            </div>
        </div>
    </div>
    """


def send_email_via_sendgrid(to_email: str, cc_emails: list[str], subject: str, html_content: str) -> tuple[bool, str]:
    sendgrid_secrets = get_secret_section("sendgrid")
    api_key = sendgrid_secrets.get("api_key")
    from_email = sendgrid_secrets.get("from_email")
    from_name = sendgrid_secrets.get("from_name", "Ferreinox S.A.S. BIC")
    if not api_key or not from_email:
        return False, "Credenciales SendGrid incompletas en st.secrets"

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "personalizations": [{"to": [{"email": to_email}], "cc": [{"email": email} for email in cc_emails if email]}],
        "from": {"email": from_email, "name": from_name},
        "subject": subject,
        "content": [{"type": "text/html", "value": html_content}],
    }
    response = requests.post("https://api.sendgrid.com/v3/mail/send", headers=headers, json=payload, timeout=30)
    if 200 <= response.status_code < 300:
        return True, f"SendGrid {response.status_code}"
    return False, f"SendGrid {response.status_code}: {response.text[:400]}"


def create_payment_lot(selected_df: pd.DataFrame, payment_date: date, responsible: str, email_destino: str) -> pd.DataFrame:
    lote_id = f"LTP-{datetime.now(COLOMBIA_TZ).strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
    lot_df = selected_df.copy()
    lot_df["lote_id"] = lote_id
    lot_df["fecha_registro"] = pd.Timestamp.now(tz=COLOMBIA_TZ)
    lot_df["fecha_programada_pago"] = pd.Timestamp(payment_date).tz_localize(COLOMBIA_TZ)
    lot_df["responsable"] = responsible
    lot_df["estado_lote"] = "Programado"
    lot_df["email_destino"] = email_destino
    lot_df.rename(columns={"valor_erp": "valor_factura", "motivo_pago": "motivo_pago"}, inplace=True)
    lot_df = lot_df[[
        "lote_id",
        "fecha_registro",
        "fecha_programada_pago",
        "responsable",
        "invoice_key",
        "proveedor",
        "num_factura",
        "valor_factura",
        "valor_descuento",
        "valor_a_pagar",
        "estado_lote",
        "motivo_pago",
        "email_destino",
    ]]
    return lot_df


def register_payment_lot(client: gspread.Client, lot_df: pd.DataFrame) -> bool:
    return append_df_to_sheet(client, SHEET_PAYMENT_LOTS, lot_df, PAYMENT_LOT_COLUMNS)


def register_email_log(client: gspread.Client, log_row: dict) -> bool:
    log_df = pd.DataFrame([log_row])
    return append_df_to_sheet(client, SHEET_EMAIL_LOG, log_df, EMAIL_LOG_COLUMNS)


def build_email_log_row(lote_id: str, provider_name: str, to_email: str, cc_email: str, subject: str, lot_df: pd.DataFrame, status: str, detail: str) -> dict:
    return {
        "envio_id": f"ENV-{uuid.uuid4().hex[:10].upper()}",
        "fecha_envio": pd.Timestamp.now(tz=COLOMBIA_TZ),
        "lote_id": lote_id,
        "proveedor": provider_name,
        "email_destino": to_email,
        "email_cc": cc_email,
        "asunto": subject,
        "facturas": ", ".join(lot_df["num_factura"].astype(str).tolist()),
        "ahorro_total": lot_df["valor_descuento"].sum(),
        "estado_envio": status,
        "detalle_envio": detail,
    }


def export_df_to_excel(df: pd.DataFrame, sheet_name: str = "Datos", title: str = "Reporte Ferreinox") -> io.BytesIO:
    """Export a DataFrame to a professionally formatted Excel file returned as BytesIO."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Sanitize values that openpyxl / pandas cannot serialize to Excel
    clean = df.copy()

    def normalize_excel_value(value: Any) -> Any:
        if value is None or pd.isna(value):
            return None
        if isinstance(value, pd.Timestamp):
            if value.tzinfo is not None:
                return value.tz_localize(None)
            return value
        if isinstance(value, datetime):
            if value.tzinfo is not None:
                return value.replace(tzinfo=None)
            return value
        return value

    for col in clean.columns:
        if pd.api.types.is_datetime64_any_dtype(clean[col]):
            series = clean[col]
            if getattr(series.dt, "tz", None) is not None:
                series = series.dt.tz_localize(None)
            clean[col] = series.where(series.notna(), other=None)
        elif pd.api.types.is_float_dtype(clean[col]) or pd.api.types.is_integer_dtype(clean[col]):
            clean[col] = clean[col].where(clean[col].notna() & np.isfinite(clean[col].astype(float)), other=0)
        else:
            clean[col] = clean[col].map(normalize_excel_value).fillna("")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        clean.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="0C2D57")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(df.columns), 1))
        sub_cell = ws.cell(row=2, column=1, value=f"Generado: {datetime.now(COLOMBIA_TZ).strftime('%Y-%m-%d %H:%M')}")
        sub_cell.font = Font(name="Calibri", size=9, italic=True, color="506070")

        # Header style
        header_fill = PatternFill(start_color="0C2D57", end_color="0C2D57", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D0D8E0"),
            right=Side(style="thin", color="D0D8E0"),
            top=Side(style="thin", color="D0D8E0"),
            bottom=Side(style="thin", color="D0D8E0"),
        )

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Data styles
        currency_cols = {c for c in df.columns if any(k in c.lower() for k in ["valor", "descuento", "ahorro", "diferencia"])}
        pct_cols = {c for c in df.columns if "pct" in c.lower() or "porcentaje" in c.lower()}
        stripe_fill = PatternFill(start_color="F5F8FB", end_color="F5F8FB", fill_type="solid")
        data_font = Font(name="Calibri", size=10, color="223548")

        for row_idx in range(4, 4 + len(df)):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if (row_idx - 4) % 2 == 1:
                    cell.fill = stripe_fill
                col_name = df.columns[col_idx - 1]
                if col_name in currency_cols:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in pct_cols:
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-width
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if not df.empty else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 12), 40)

        ws.sheet_properties.tabColor = "0C2D57"
        ws.freeze_panes = "A4"

    buf.seek(0)
    return buf


def get_discount_summary_for_suppliers() -> pd.DataFrame:
    """Build a summary table of all configured supplier discount rules."""
    rows = []
    for supplier, rules in DISCOUNT_PROVIDERS.items():
        for rule in sorted(rules, key=lambda r: r["days"]):
            rows.append({
                "Proveedor": supplier,
                "Días límite": rule["days"],
                "Descuento %": rule["rate"],
            })
    return pd.DataFrame(rows)
